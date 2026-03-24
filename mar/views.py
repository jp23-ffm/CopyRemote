import csv
import datetime
import django
import io
import json
import math
import os
import time
import threading
import uuid
import socket

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.cache import cache
from django.db import models, transaction
from django.db.models import Q, F, Case, When, Count, OuterRef, Subquery
from django.db.models.functions import Upper
from django.http import FileResponse, JsonResponse, HttpResponse, HttpResponseRedirect, StreamingHttpResponse, QueryDict
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

from collections import defaultdict
from .exports import generate_csv, generate_excel, EXPORT_DIR
from urllib.parse import urlencode, parse_qs, unquote
from functools import lru_cache
from threading import Lock

from common.views import generate_charts
from .models import AnalysisSnapshot, ServerDiscrepancy, DiscrepancyTracking, DiscrepancyAnnotation, ImportStatus, ExcludedServer
from .utils import get_trend_data
from userapp.models import UserProfile, SavedSearch, SavedOptions, UserPermissions


app_name=__package__.split('.')[-1]

_field_labels_cache = None
_field_labels_file_mtime = 0
_cache_lock = Lock()


def get_field_labels():
    # Cache field_labels.json with update when the file changes, Thread-safe for Gunicorn

    global _field_labels_cache, _field_labels_file_mtime
    
    json_path = os.path.join(os.path.dirname(__file__), 'field_labels.json')
    
    try:
        current_file_mtime = os.path.getmtime(json_path)
    except OSError:
        return _field_labels_cache

    if _field_labels_cache is not None and current_file_mtime == _field_labels_file_mtime:
        return _field_labels_cache
    
    with _cache_lock:
        try:
            current_file_mtime = os.path.getmtime(json_path)
        except OSError:
            return _field_labels_cache
        
        if _field_labels_cache is not None and current_file_mtime == _field_labels_file_mtime:
            return _field_labels_cache
        
        try:
            with open(json_path, 'r', encoding="utf-8") as f:
                _field_labels_cache = json.load(f)
            
            _field_labels_file_mtime = current_file_mtime
            
            #print(f"[get_field_labels] Reloaded field_labels.json (mtime: {current_file_mtime})")
            
        except (OSError, json.JSONDecodeError) as e:
            #print(f"[get_field_labels] Error loading field_labels.json: {e}")
            return _field_labels_cache
    
    return _field_labels_cache


def _compute_days_open(tracker, filtered_fields=None):
    """
    Return days_open for a tracker instance.
    If filtered_fields is provided and has matching active issues, return the
    max days_open among those fields (context-sensitive).
    Otherwise return the global max (oldest_first_seen).
    """
    if not tracker or not tracker.active_issues:
        return ''

    now = timezone.now()

    if filtered_fields:
        relevant = {f: v for f, v in tracker.active_issues.items() if f in filtered_fields}
        if relevant:
            return max(
                (now - parse_datetime(v['first_seen'])).days + 1
                for v in relevant.values()
            )

    if tracker.oldest_first_seen:
        return (now - tracker.oldest_first_seen).days + 1

    return max(
        (now - parse_datetime(v['first_seen'])).days + 1
        for v in tracker.active_issues.values()
    )


def get_filter_mapping():

    json_path = os.path.join(os.path.dirname(__file__), 'field_labels.json')
    with open(json_path, "r", encoding="utf-8") as f:
        field_labels = json.load(f)

    # Initialize an empty dictionary to store the filter mapping
    FILTER_MAPPING = {}
    
    # Iterate over each field defined in the fields section of the JSON configuration
    for field_name, field_properties in field_labels['fields'].items():
        FILTER_MAPPING[field_name] = field_name  # Map the field to the attribute of the model

    # Return the filter mapping dictionary
    return FILTER_MAPPING
    
    
def construct_query(key, terms):
    # Creates a Django Q object based on a list of terms for a specific field
    query = Q()
    
    # Iterate over each term in the terms list
    for term in terms:
        if term.startswith('@'):  # Check if the term starts with '@' for an exact match
            term = term[1:]  # Remove the '@' character
            query |= Q(**{f'{key}__iexact': term})  # Create a Q object for an exact match (case-insensitive)
        elif term.startswith('!'):  # Check if the term starts with '!' for an exclusion
            term = term[1:]  # Remove the '!' character
            query &= ~Q(**{f'{key}__icontains': term})  # Create a Q object for an exclusion (case-insensitive containment test)
        else:  # For terms without special characters, perform a containment test
            query |= Q(**{f'{key}__icontains': term})  # Create a Q object for a containment test (case-insensitive)

    # Return the combined Q object representing the filter criteria
    return query
       

# View to display the server information - Main View
@login_required
def server_view(request):

    # Get the user's profile if it already exists in the table userapp_userprofile
    localhostname = socket.gethostname()    

    try:
        profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        profile = UserProfile.objects.create(user=request.user)

    # Read and save the field_labels.json information
    json_data = get_field_labels()
     
    # Initialize the filters dictionary
    filters = {}

    # Create the filter using the parameter passed in the URL with request.GET.get
    for field_key, field_info in json_data['fields'].items():
        input_name = field_info.get('inputname')
        if input_name:
            filter_value = request.GET.get(input_name, '').split(',')
            filters[field_key] = [v for v in filter_value if v]

    # Build the query based on the filters values extracted earlier
    all_servers = ServerDiscrepancy.objects.all()

    excluded_names = set(ExcludedServer.objects.values_list('server_name', flat=True))
    if excluded_names:
        all_servers = all_servers.exclude(SERVER_ID__in=excluded_names)

    combined_filter_query = Q()
    for key, values in filters.items():
        if not values:
            continue

        # Build the query for this field
        query = construct_query(key, values)
        combined_filter_query &= query  # Combine with AND

    if combined_filter_query:
        all_servers = all_servers.filter(combined_filter_query)

    # any_inconsistency=KO → OR between alive and dead inconsistency (virtual param)
    any_inc = request.GET.get('any_inconsistency', '').strip()
    if any_inc:
        all_servers = all_servers.filter(
            Q(alive_status_inconsistent__icontains=any_inc) | Q(dead_status_inconsistent__icontains=any_inc)
        )

    # days_open filter — restrict to servers open for at least N days
    days_open_param = request.GET.get('days_open', '').strip()
    if days_open_param.isdigit() and int(days_open_param) > 0:
        cutoff = timezone.now() - datetime.timedelta(days=int(days_open_param))
        matching_ids = DiscrepancyTracking.objects.filter(
            oldest_first_seen__lte=cutoff
        ).values_list('SERVER_ID', flat=True)
        all_servers = all_servers.filter(SERVER_ID__in=matching_ids)

    # Get filtered and sorted servers
    sort_field = request.GET.get('sort', 'SERVER_ID')
    sort_order = request.GET.get('order', 'asc')

    if sort_field == 'days_open':
        oldest = DiscrepancyTracking.objects.filter(
            SERVER_ID=OuterRef('SERVER_ID')
        ).values('oldest_first_seen')[:1]
        all_servers = all_servers.annotate(oldest_issue=Subquery(oldest))
        order_expr = F('oldest_issue').asc(nulls_last=True) if sort_order == 'asc' else F('oldest_issue').desc(nulls_last=True)
    else:
        # Validate sort_field against known fields to prevent injection
        valid_sort_fields = {f_key for f_key in json_data.get('fields', {}).keys()}
        if sort_field not in valid_sort_fields:
            sort_field = 'SERVER_ID'
        order_expr = F(sort_field).asc(nulls_last=True) if sort_order == 'asc' else F(sort_field).desc(nulls_last=True)

    filtered_servers = all_servers.order_by(order_expr)
   
    # Define some pagination settings that will be passed as context
    page_size = int(request.GET.get('page_size', 50))
    page_number = request.GET.get("page")
   
    # Json part - Reading of field_labels.json to set element dynamically
    
    # Read and store the sections categories and fields of field_labels.json
    json_data_categories=json_data.get("categories", {})
    json_data_fields=json_data.get("fields", {})
    validation_errors = json_data.get('validation_errors', {})

    # Generation of the information loaded form field_labels.json to create the search boxes, the dropdown lists and populate their content
    finalfields = [(field, info) for field, info in json_data_fields.items()]  # Read all items in the json fields section
    
    table_fields=[]
    # Loop in the fields items: if some have the attribute listbox, generate the content to display it the associated drop down list in the view

    cacheset=False
    for key, val in finalfields:
        listbox_value=val.get("listbox", '')
        listempty_value=val.get("listempty", '') 
        if listbox_value:
            cache_key = f"listbox_disc_{key}"
            listbox_evaluated_disc = cache.get(cache_key)
            if listbox_evaluated_disc is None:
                if listempty_value =="True":
                    listbox_evaluated_disc = ['MISSING']
                else:
                    listbox_evaluated_disc = ServerDiscrepancy.objects.values_list(key, flat=True).distinct().order_by(key)

                if listbox_evaluated_disc:  # Sort the entry to put "EMPTY" at the end
                    listbox_evaluated_disc = list(listbox_evaluated_disc)
                    if any(x is None or x.upper() == "EMPTY" for x in listbox_evaluated_disc):
                        has_na = any(isinstance(x, str) and x.upper() == "EMPTY" for x in listbox_evaluated_disc)  #"EMPTY" in listbox_evaluated_disc
                        listbox_evaluated_disc = [
                            x for x in listbox_evaluated_disc if x is not None and x != "" and x.upper() != "EMPTY"
                        ]

                        listbox_evaluated_disc.sort()
                        if has_na:
                            listbox_evaluated_disc.append("EMPTY")
                   
                cache.set(cache_key, listbox_evaluated_disc, timeout=3600)  # Cache for 1 hour
                cacheset=True
        else:
            listbox_evaluated_disc = '' 
          
        # Create a table_fields item with information from the json: this information will be used for the HTML creation
        table_fields.append({
    		"name":key,
    	  	"displayname":val.get("displayname", key),
    	  	"inputname":val.get("inputname", key),
    		"listbox":listbox_evaluated_disc,
    	  	"listboxmsg":val.get("listboxmsg", 'Select an option'),
    	  	"listid":val.get("listid", 'missingid'),
    	  	"selectionsection":val.get("selectionsection", 'cat0'),
        })
        
    # Generation of the information loaded form field_labels.json to create the category tree and its sub-categories
    grouped=defaultdict(list)
    for key, value in json_data_fields.items():
        if (isinstance(value, dict)):
            section=value.get('selectionsection', '').strip()
            displayname=value.get('displayname', '').strip()
            ischecked=bool(value.get('ischecked') == "True")
            ischeckeddisabled=bool(value.get('ischeckeddisabled') == "True")
            
            if not displayname:
               displayname=key

            if section in json_data_categories:
                grouped[section].append({
                    'key':key,
                    'displayname':displayname,
                    'ischecked':ischecked,
                    'ischeckeddisabled':ischeckeddisabled
                })
                
    category_fields = [
        {
            'category':cat,
            'title':json_data_categories[cat],
            'fields': grouped[cat]
        }
        for cat in json_data_categories
        if cat in grouped
    ]
  
    profile = UserProfile.objects.get(user=request.user)
    last_status = ImportStatus.objects.order_by('-date_import').first()
    visible_columns = request.GET.get("visible_columns")
    
    # Processing according to display mode
    display_servers = []
    
    # Get model fields (exclude technical fields)
    model_fields = []
    excluded_fields = {field_name for field_name in json_data['fields'].keys() 
        if field_name.endswith('_inconsistent')}

    for field_name, field_info in json_data['fields'].items():     
        if field_name not in excluded_fields:
            model_fields.append({
                'name': field_name,
                'verbose_name': field_info.get('displayname', field_name.replace('_', ' ').title()),
                'is_hostname': field_name == 'SERVER_ID'
            })

    
    paginator = Paginator(filtered_servers, page_size)
    page_obj_raw = paginator.get_page(request.GET.get('page'))

    # Convert only the current page to list
    current_page_servers = list(page_obj_raw)

    # Process servers from current page
    for server in current_page_servers:
        display_servers.append({
            'hostname': server.SERVER_ID,
            'primary_server': server
        })

    # Fetch tracker data for current page servers
    hostnames_in_page = [s['hostname'] for s in display_servers]
    if hostnames_in_page:
        trackers = DiscrepancyTracking.objects.filter(SERVER_ID__in=hostnames_in_page)
        tracker_dict = {t.SERVER_ID: t for t in trackers}

        # Fields actively filtered — used for context-sensitive days_open
        active_filter_fields = {k for k, v in filters.items() if v and k not in ('days_open', 'ANNOTATION')}

        for server_group in display_servers:
            tracker = tracker_dict.get(server_group['hostname'])
            server_group['tracker'] = tracker
            server_group['tracker_json'] = json.dumps(tracker.active_issues) if tracker else '{}'
            server_group['primary_server'].days_open = _compute_days_open(tracker, active_filter_fields)


    # Load annotations for current page servers
    annotations = DiscrepancyAnnotation.objects.filter(SERVER_ID__in=hostnames_in_page)
    annotations_dict = {ann.SERVER_ID: ann for ann in annotations}
    for server_group in display_servers:
        server_group['annotation'] = annotations_dict.get(server_group['hostname'])
    
    page_obj = create_page_wrapper(display_servers, page_obj_raw)
    total_servers_stat = paginator.count

    # Load saved searches for this user/app
    saved_searches = SavedSearch.objects.filter(
        user_profile__user=request.user, view=app_name
    ).order_by('name')

    # Rendering servers.html with the corresponding context

    context = {
        'page_obj' : page_obj,  # Pagination object
        'table_fields' : table_fields,  # Information for the dynamic creation of table columns, input boxes and list boxes
        'category_fields' : category_fields,  # Information for the dynamic creation of the catgeory tree
        'appname' : app_name,  # Application name
        'page_size' : page_size,  # Number of objects per page
        'current_filters': filters,  # Filters defined above
        'json_data' : json.dumps(json_data),  # Json content
        'last_status' : last_status,  # Last import state
        'loggedonuser' : request.user,  # Logon name
        'model_fields': model_fields,
        'total_servers': total_servers_stat,
        'cacheset': cacheset,
        'sort_field': sort_field,
        'sort_order': sort_order,
        'localhostname': localhostname,
        'validation_errors': validation_errors,
        'saved_searches': saved_searches,
    }

    return render(request, f'{app_name}/servers.html', context)


@login_required
def dashboard_view(request):
    # Main dashboard view displaying data quality metrics

    # Load dashboard configuration from JSON
    config_path = os.path.join(os.path.dirname(__file__), 'discrepancies_dashboard.json' )
    with open(config_path, 'r') as f:
        config = json.load(f)

    # Permanent filter — read saved selection for this user
    json_data = get_field_labels()
    permanent_filter_names = list(json_data.get('permanentfilters', {}).keys())
    permanent_filter_selection = 'All Servers'
    permanent_filter_label = 'All Servers'
    permanent_filter_description = ''
    saved_searches = SavedSearch.objects.none()
    fields_info = json_data.get('fields', {})
    field_to_inputname = {fname: finfo.get('inputname', fname.lower())
                          for fname, finfo in fields_info.items()}
    inputname_to_field = {finfo.get('inputname', fname.lower()): fname
                          for fname, finfo in fields_info.items()}
    field_displayname   = {fname: finfo.get('displayname', fname)
                          for fname, finfo in fields_info.items()}
    if request.user.is_authenticated:
        try:
            profile = UserProfile.objects.get(user=request.user)
            saved_searches = list(SavedSearch.objects.filter(
                user_profile=profile, view=app_name
            ).order_by('name'))
            user_options = SavedOptions.objects.get(user_profile=profile)
            saved = user_options.discrepancies_permanentfilter
            if saved and saved in permanent_filter_names:
                permanent_filter_selection = saved
                permanent_filter_label = saved
                pf_def = json_data.get('permanentfilters', {}).get(saved, {})
                permanent_filter_description = ' & '.join(
                    f'{field_displayname.get(f, f)} = {", ".join(val.lstrip("@!") for val in v)}'
                    for f, v in pf_def.items()
                )
            elif saved and saved.startswith('ss:'):
                try:
                    ss_id = int(saved[3:])
                    ss = SavedSearch.objects.get(id=ss_id, user_profile=profile)
                    permanent_filter_selection = saved
                    permanent_filter_label = ss.name
                    _METRIC_FIELDS = {'alive_status_inconsistent', 'dead_status_inconsistent', 'missing_fields'}
                    parts = []
                    for inputname, value in ss.filters.items():
                        field = inputname_to_field.get(inputname, inputname)
                        if field in _METRIC_FIELDS:
                            continue
                        label = field_displayname.get(field, field)
                        clean = [v.lstrip('@!') for v in (value.split(',') if isinstance(value, str) else value)]
                        parts.append(f'{label} = {", ".join(clean)}')
                    permanent_filter_description = ' & '.join(parts) if parts else ss.name
                except (ValueError, SavedSearch.DoesNotExist):
                    pass
            elif saved and saved.startswith('multi:'):
                permanent_filter_selection = saved
                _METRIC_FIELDS = {'alive_status_inconsistent', 'dead_status_inconsistent', 'missing_fields'}
                desc_parts = []
                label_parts = []
                for part in saved[6:].split('|'):
                    if ':' not in part:
                        continue
                    inputname, raw_values = part.split(':', 1)
                    field = inputname_to_field.get(inputname)
                    if field and field not in _METRIC_FIELDS:
                        lbl = field_displayname.get(field, field)
                        clean = [v.lstrip('@!') for v in raw_values.split(',') if v.strip()]
                        desc_parts.append(f'{lbl} = {", ".join(clean)}')
                        label_parts.extend(clean)
                permanent_filter_description = ' & '.join(desc_parts)
                permanent_filter_label = ', '.join(label_parts) if label_parts else 'Multi-filter'
        except (UserProfile.DoesNotExist, SavedOptions.DoesNotExist):
            pass

    # Get the latest analysis snapshot
    try:
        latest_snapshot = AnalysisSnapshot.objects.latest('analysis_date')
    except AnalysisSnapshot.DoesNotExist:
        # No analysis has been run yet
        context = {
            'no_data': True,
            'message': 'No analysis has been run yet. Please run the discrepancy analysis first.',
        }
        return render(request, 'discrepancies/discrepancies_dashboard.html', context)

    # ── Mapping metric name → field name (for missing_fields__icontains) ───
    _METRIC_TO_FIELD = {
        'missing_live_status_count':       'LIVE_STATUS',
        'missing_osshortname_count':       'OSSHORTNAME',
        'missing_osfamily_count':          'OSFAMILY',
        'missing_snow_supportgroup_count': 'SNOW_SUPPORTGROUP',
        'missing_machine_type_count':      'MACHINE_TYPE',
        'missing_manufacturer_count':      'MANUFACTURER',
        'missing_country_count':           'COUNTRY',
        'missing_app_auid_value_count':    'APP_AUID_VALUE',
        'missing_app_name_value_count':    'APP_NAME_VALUE',
        'missing_region_count':            'REGION',
        'missing_city_count':              'CITY',
        'missing_infraversion_count':      'INFRAVERSION',
        'missing_ipaddress_count':         'IPADDRESS',
        'missing_snow_status_count':       'SNOW_STATUS',
        'missing_idrac_name_count':        'IDRAC_NAME',
        'missing_idrac_ip_count':          'IDRAC_IP',
    }

    # ── Live hero stats (always computed, never from snapshot) ──────────────
    from inventory.models import Server as InventoryServer
    _pf_defs = json_data.get('permanentfilters', {}).values()
    pf_region_values = sorted(set(v for d in _pf_defs for v in d.get('REGION', [])))
    pf_os_values     = sorted(set(v for d in _pf_defs for v in d.get('OSFAMILY', [])))
    _excluded = set(ExcludedServer.objects.values_list('server_name', flat=True))
    _disc_base = ServerDiscrepancy.objects.all()
    if _excluded:
        _disc_base = _disc_base.exclude(SERVER_ID__in=_excluded)

    _inv_eligible = InventoryServer.objects.filter(
        LIVE_STATUS='ALIVE', SNOW_STATUS='OPERATIONAL', INFRAVERSION__in=['IV1', 'IV2', 'IBM'])
    if _excluded:
        _inv_eligible = _inv_eligible.exclude(SERVER_ID__in=_excluded)
    _total_eligible = _inv_eligible.values('SERVER_ID').distinct().count()
    _total_physical = _inv_eligible.filter(MACHINE_TYPE='PHYSICAL').values('SERVER_ID').distinct().count()

    _servers_with_missing = _disc_base.filter(missing_fields__gt='').count()
    _servers_clean_missing = max(0, _total_eligible - _servers_with_missing)
    _pct_missing = math.trunc((_servers_clean_missing / _total_eligible * 10000)) / 100 if _total_eligible else 100

    _inv_all = InventoryServer.objects.all()
    if _excluded:
        _inv_all = _inv_all.exclude(SERVER_ID__in=_excluded)
    _total_all = _inv_all.values('SERVER_ID').distinct().count()
    _inc_count = _disc_base.filter(
        Q(alive_status_inconsistent='KO') | Q(dead_status_inconsistent='KO')
    ).count()
    _servers_ok_inc = max(0, _total_all - _inc_count)
    _pct_inc_ok = math.trunc((_servers_ok_inc / _total_all * 10000)) / 100 if _total_all else 100

    _hero_stats = {
        'missing_data': {
            'value': _pct_missing,
            'servers_with_issues': _servers_with_missing,
            'servers_clean': _servers_clean_missing,
            'total_servers': _total_eligible,
            'total_physical': _total_physical,
        },
        'operational_inconsistencies': {
            'value': _pct_inc_ok,
            'servers_with_issues': _inc_count,
            'servers_clean': _servers_ok_inc,
            'total_servers': _total_all,
            'total_physical': _total_physical,
        },
    }
    # ────────────────────────────────────────────────────────────────────────

    # Prepare widget data
    widgets_data = []

    existing_trackers = {t.SERVER_ID: t for t in DiscrepancyTracking.objects.all()}

    for widget in config['dashboard']['widgets']:
        widget_data = {
            'id': widget['id'],
            'type': widget['type'],
            'size': widget.get('size', 'small'),
            'title': widget['title'],
            'icon': widget.get('icon', ''),
            'link': widget.get('link', '#'),
            'info': widget.get('info', ''),
            'metric': widget.get('metric', ''),
            'physical_only': widget.get('physical_only', 'no') == 'yes',
        }
        
        metric_name = widget['metric']

        if widget['type'] == 'gauge':
            if widget['size'] == 'large':
                # Large gauge: always computed live
                hs = _hero_stats.get(widget['id'], {})
                widget_data['value'] = hs.get('value', 0)
                widget_data['display_value'] = f"{hs.get('value', 0)}%"
                widget_data['servers_with_issues'] = hs.get('servers_with_issues', 0)
                widget_data['servers_clean'] = hs.get('servers_clean', 0)
                widget_data['total_servers'] = hs.get('total_servers', 0)
            else:
                # Small gauge: computed live from _disc_base / _inv_eligible
                if metric_name == 'alive_status_inconsistent_count':
                    metric_value = _disc_base.filter(alive_status_inconsistent='KO').count()
                elif metric_name == 'dead_status_inconsistent_count':
                    metric_value = _disc_base.filter(dead_status_inconsistent='KO').count()
                else:
                    # missing_<field>_count → field name via FIELD_TO_METRIC reverse
                    field = _METRIC_TO_FIELD.get(metric_name)
                    metric_value = _disc_base.filter(missing_fields__icontains=field).count() if field else 0

                total = _total_physical if widget.get('physical_only', 'no') == 'yes' else _total_eligible
                servers_ok = max(0, total - metric_value)
                percentage_ok = math.trunc((servers_ok / total) * 10000) / 100 if total else 100

                widget_data['value'] = percentage_ok
                widget_data['count'] = metric_value
                widget_data['display_value'] = f"{percentage_ok}%"
                widget_data['detail'] = f"{metric_value} / {total} issues"
                widget_data['servers_clean'] = servers_ok
                widget_data['servers_with_issues'] = metric_value
                widget_data['total_servers'] = total
            
            # Thresholds and colors for gauge coloring
            widget_data['thresholds'] = widget.get('thresholds', {
                'critical': 80,
                'warning': 95,
                'good': 100
            })
            widget_data['colors'] = widget.get('colors', {
                'critical': '#dc3545',
                'warning': '#ffc107',
                'good': '#28a744'
            })
        
        widgets_data.append(widget_data)
    
    # Prepare historical trend data (default metric)
    historic_config = config['dashboard'].get('historic_section', {})
    
    if historic_config.get('enabled', False):
        default_metric = historic_config.get('default_metric', 'servers_with_issues')
        days = historic_config.get('days', 30)
        trend_data = get_trend_data(default_metric, days)
    else:
        historic_config = None
        trend_data = None

    recent_snapshots = list(AnalysisSnapshot.objects.order_by('-analysis_date')[:10])

    raw_diff = latest_snapshot.diff_summary or {}
    diff_new      = raw_diff.get('new', [])
    diff_resolved = raw_diff.get('resolved', [])
    diff_changed  = raw_diff.get('changed', {})
    SHOW_MAX = 20
    diff_display = {
        'is_first_run':    not raw_diff,
        'new':             diff_new[:SHOW_MAX],
        'new_total':       len(diff_new),
        'new_more':        max(0, len(diff_new) - SHOW_MAX),
        'resolved':        diff_resolved[:SHOW_MAX],
        'resolved_total':  len(diff_resolved),
        'resolved_more':   max(0, len(diff_resolved) - SHOW_MAX),
        'changed':         list(diff_changed.items())[:SHOW_MAX],
        'changed_total':   len(diff_changed),
        'changed_more':    max(0, len(diff_changed) - SHOW_MAX),
    }
    
    context = {
        'title': config['dashboard']['title'],
        'widgets': widgets_data,
        'snapshot': latest_snapshot,
        'no_data': False,
        'has_exclusions': len(_excluded) > 0,
        'exclusions_count': len(_excluded),
        'historic_config': historic_config,
        'trend_data': trend_data,
        'recent_snapshots': recent_snapshots,
        'diff_display': diff_display,
        'permanent_filter_names': permanent_filter_names,
        'permanent_filter_selection': permanent_filter_selection,
        'permanent_filter_label': permanent_filter_label,
        'permanent_filter_description': permanent_filter_description,
        'saved_searches': saved_searches,
        'pf_region_values': pf_region_values,
        'pf_os_values': pf_os_values,
    }

    return render(request, 'discrepancies/discrepancies_dashboard.html', context)


@login_required
def dashboard_filter_api(request):
    """
    AJAX endpoint — returns widget counts for a given permanent filter selection.
    GET ?pf=Windows  →  JSON with total_servers, total_physical, metrics{...}, filter_link_params
    If pf is empty or 'All Servers', returns data from the latest AnalysisSnapshot.
    Otherwise computes live counts from ServerDiscrepancy filtered by the selection.
    """
    FIELD_TO_METRIC = {
        'LIVE_STATUS':       'missing_live_status_count',
        'OSSHORTNAME':       'missing_osshortname_count',
        'OSFAMILY':          'missing_osfamily_count',
        'SNOW_SUPPORTGROUP': 'missing_snow_supportgroup_count',
        'MACHINE_TYPE':      'missing_machine_type_count',
        'MANUFACTURER':      'missing_manufacturer_count',
        'COUNTRY':           'missing_country_count',
        'APP_AUID_VALUE':    'missing_app_auid_value_count',
        'APP_NAME_VALUE':    'missing_app_name_value_count',
        'REGION':            'missing_region_count',
        'CITY':              'missing_city_count',
        'INFRAVERSION':      'missing_infraversion_count',
        'IPADDRESS':         'missing_ipaddress_count',
        'SNOW_STATUS':       'missing_snow_status_count',
        'IDRAC_NAME':        'missing_idrac_name_count',
        'IDRAC_IP':          'missing_idrac_ip_count',
    }

    filter_name = request.GET.get('pf', '').strip()
    json_data = get_field_labels()
    permanent_filters = json_data.get('permanentfilters', {})
    fields_info = json_data.get('fields', {})
    field_to_inputname = {fname: finfo.get('inputname', fname.lower())
                          for fname, finfo in fields_info.items()}
    inputname_to_field = {finfo.get('inputname', fname.lower()): fname
                          for fname, finfo in fields_info.items()}
    field_displayname  = {fname: finfo.get('displayname', fname)
                          for fname, finfo in fields_info.items()}

    from inventory.models import Server as InventoryServer

    # Fields that are dashboard metrics themselves — meaningless as population filters
    DASHBOARD_METRIC_FIELDS = {'alive_status_inconsistent', 'dead_status_inconsistent', 'missing_fields'}

    # Resolve filter_def — handles both system permanent filters and ss: saved-search entries
    filter_def = {}
    if filter_name and filter_name in permanent_filters:
        filter_def = permanent_filters[filter_name]
    elif filter_name and filter_name.startswith('ss:'):
        try:
            ss_id = int(filter_name[3:])
            ss = SavedSearch.objects.get(id=ss_id, user_profile__user=request.user)
            for inputname, value in ss.filters.items():
                field = inputname_to_field.get(inputname)
                if field and field not in DASHBOARD_METRIC_FIELDS:
                    filter_def[field] = (
                        [v.strip() for v in value.split(',') if v.strip()]
                        if isinstance(value, str) else value
                    )
        except (ValueError, SavedSearch.DoesNotExist):
            pass
    elif filter_name and filter_name.startswith('multi:'):
        for part in filter_name[6:].split('|'):
            if ':' not in part:
                continue
            inputname, raw_values = part.split(':', 1)
            field = inputname_to_field.get(inputname)
            if field and field not in DASHBOARD_METRIC_FIELDS:
                values = [v.strip() for v in raw_values.split(',') if v.strip()]
                if values:
                    filter_def[field] = values

    is_filtered = bool(filter_def)
    disc_q = Q()
    inv_q  = Q()
    link_parts = []
    filter_parts = []   # human-readable description pieces
    if is_filtered:
        inv_field_names = {f.name for f in InventoryServer._meta.get_fields()}
        for field, values in filter_def.items():
            disc_q &= construct_query(field, values)
            if field in inv_field_names:
                inv_q &= construct_query(field, values)
            inputname = field_to_inputname.get(field, field.lower())
            link_parts.append(f'{inputname}={",".join(values)}')
            # Build readable label: strip @/! prefixes, use displayname
            clean_values = [v.lstrip('@!') for v in values]
            label = field_displayname.get(field, field)
            filter_parts.append(f'{label} = {", ".join(clean_values)}')

    excluded_names = set(ExcludedServer.objects.values_list('server_name', flat=True))

    disc_qs = ServerDiscrepancy.objects.filter(disc_q) if disc_q else ServerDiscrepancy.objects.all()
    if excluded_names:
        disc_qs = disc_qs.exclude(SERVER_ID__in=excluded_names)

    # ── Days-open filter ─────────────────────────────────────────────────
    days_open_str = request.GET.get('days_open', '').strip()
    days_open_int = int(days_open_str) if days_open_str.isdigit() and int(days_open_str) > 0 else 0
    if days_open_int > 0:
        cutoff = timezone.now() - datetime.timedelta(days=days_open_int)
        active_ids = DiscrepancyTracking.objects.filter(
            oldest_first_seen__lte=cutoff
        ).values_list('SERVER_ID', flat=True)
        disc_qs = disc_qs.filter(SERVER_ID__in=active_ids)
        link_parts.append(f'days_open={days_open_int}')

    # ── Missing Data hero gauge ──────────────────────────────────────────
    base_eligible_q = Q(LIVE_STATUS='ALIVE', SNOW_STATUS='OPERATIONAL', INFRAVERSION__in=['IV1', 'IV2', 'IBM'])
    inv_eligible = InventoryServer.objects.filter(base_eligible_q & inv_q) if inv_q else InventoryServer.objects.filter(base_eligible_q)
    if excluded_names:
        inv_eligible = inv_eligible.exclude(SERVER_ID__in=excluded_names)
    total_eligible = inv_eligible.values('SERVER_ID').distinct().count()
    total_physical = inv_eligible.filter(MACHINE_TYPE='PHYSICAL').values('SERVER_ID').distinct().count()

    servers_with_missing = disc_qs.filter(missing_fields__gt='').count()
    servers_clean_missing = max(0, total_eligible - servers_with_missing)
    pct_missing = math.trunc((servers_clean_missing / total_eligible) * 10000) / 100 if total_eligible else 100

    # ── Operational Inconsistencies hero gauge ───────────────────────────
    inv_all = InventoryServer.objects.filter(inv_q) if inv_q else InventoryServer.objects.all()
    if excluded_names:
        inv_all = inv_all.exclude(SERVER_ID__in=excluded_names)
    total_all = inv_all.values('SERVER_ID').distinct().count()

    inc_count = disc_qs.filter(
        Q(alive_status_inconsistent='KO') | Q(dead_status_inconsistent='KO')
    ).count()
    servers_ok_inc = max(0, total_all - inc_count)
    pct_inc_ok = math.trunc((servers_ok_inc / total_all) * 10000) / 100 if total_all else 100

    # ── Small gauge metrics ───────────────────────────────────────────────
    metrics = {}
    for field, metric_name in FIELD_TO_METRIC.items():
        metrics[metric_name] = disc_qs.filter(missing_fields__icontains=field).count()
    metrics['alive_status_inconsistent_count'] = disc_qs.filter(alive_status_inconsistent='KO').count()
    metrics['dead_status_inconsistent_count']  = disc_qs.filter(dead_status_inconsistent='KO').count()

    # Metrics that are meaningless under the current filter (the filtered field can't be "missing")
    grayed_metrics = []
    if is_filtered:
        for field in filter_def:
            metric = FIELD_TO_METRIC.get(field)
            if metric:
                grayed_metrics.append(metric)

    return JsonResponse({
        # Missing Data gauge
        'total_eligible':        total_eligible,
        'total_physical':        total_physical,
        'servers_with_missing':  servers_with_missing,
        'servers_clean_missing': servers_clean_missing,
        'pct_missing_clean':     pct_missing,
        # Operational Inconsistencies gauge
        'total_all':             total_all,
        'inconsistency_count':   inc_count,
        'servers_ok_inc':        servers_ok_inc,
        'pct_inc_ok':            pct_inc_ok,
        # Small gauges
        'metrics':               metrics,
        'grayed_metrics':        grayed_metrics,
        'filter_link_params':    '&'.join(link_parts),
        'filter_description':    ' & '.join(filter_parts) if filter_parts else '',
    })


def trend_api_view(request):
    """
    API endpoint for retrieving trend data for a specific metric.
    Called via AJAX when user changes the metric selector.
    
    Query params:
        - metric: Field name from AnalysisSnapshot
        - days: Number of days to look back (default: 30)
    
    Returns:
        JSON with dates (labels) and values
    """
    
    metric = request.GET.get('metric', 'servers_with_issues')
    days = int(request.GET.get('days', 30))
    
    data = get_trend_data(metric, days)
    
    return JsonResponse({
        'labels': data['dates'],
        'values': data['values'],
        'metric': metric
    })


def create_page_wrapper(object_list, source_page):
    # Create a unified page object wrapper
    
    class UnifiedPageObj:
        def __init__(self, objects, source):
            self.object_list = objects
            self.number = source.number
            self.paginator = source.paginator
            
        def __iter__(self):
            return iter(self.object_list)
            
        def has_previous(self):
            return self.number > 1
            
        def has_next(self):
            return self.number < self.paginator.num_pages
            
        def previous_page_number(self):
            return self.number - 1 if self.has_previous() else None
            
        def next_page_number(self):
            return self.number + 1 if self.has_next() else None
            
        def has_other_pages(self):
            return self.paginator.num_pages > 1
    
    return UnifiedPageObj(object_list, source_page)


@csrf_exempt
def update_permanentfilter_field(request):
    if request.method == "POST":
        
        # Get the permanent filter choice from the POST data, passed as input permanentfilter_choice
        permanentfilter_choice = request.POST.get("permanentfilter_choice")

        if request.user.is_authenticated:
            # Get the user's profile
            profile = UserProfile.objects.get(user=request.user)

            # Update or create the SavedOptions entry for the user
            permanentfilter, created = SavedOptions.objects.update_or_create(  # Update or create the SavedOptions entry for the user
                user_profile=profile,
                defaults={f'{app_name}_permanentfilter': permanentfilter_choice}
            )
            # Redirect back to the referring page after submission
            return redirect(request.META.get('HTTP_REFERER', '/'))  # Redirect back to the same page after submission
        else:
            # The user is not authenticated
            return JsonResponse({"status": "error", "message": "User is not authenticated."}, status=401)

    # Return an error message if the request method is invalid
    return JsonResponse({"status": "error", "message": "Invalid request method."}, status=400)


@login_required
def edit_annotation(request, hostname):
    annotation = DiscrepancyAnnotation.objects.filter(SERVER_ID=hostname).first()

    if request.method == 'GET':
        return JsonResponse({
            'hostname': hostname,
            'comment': annotation.comment if annotation else '',
            'assigned_to': annotation.assigned_to if annotation else '',
            'history': annotation.get_history_display() if annotation else [],
        })

    elif request.method == 'POST':
        comment = request.POST.get('comment', '').strip()
        assigned_to = request.POST.get('assigned_to', '').strip()

        if not annotation:
            annotation = DiscrepancyAnnotation(SERVER_ID=hostname)

        annotation.add_entry(comment, assigned_to, request.user)

        return JsonResponse({
            'success': True,
            'message': 'Annotation saved successfully',
            'comment': annotation.comment,
            'assigned_to': annotation.assigned_to,
            'history': annotation.get_history_display(),
            'updated_by': request.user.username,
            'updated_at': annotation.updated_at.strftime('%d/%m/%Y %H:%M')
        })

    return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=400)


@login_required
def bulk_annotation(request):
    if request.method == "POST":
        try:
            query_param = request.POST.get('query')
            if not query_param:
                return JsonResponse({'status': 'error', 'message': 'Query string not provided'}, status=400)

            query_param = unquote(query_param)
            parsed_query = parse_qs(query_param)

            json_data = get_field_labels()

            # Build filter query from URL params
            servers_to_update = ServerDiscrepancy.objects.all()
            filters = {}
            for field_key, field_info in json_data['fields'].items():
                if 'inputname' in field_info:
                    filters[field_key] = parsed_query.get(field_info['inputname'], [''])[0].split(',')
            filters = {k: v for k, v in filters.items() if v != ['']}

            for key, value in filters.items():
                terms = value if isinstance(value, list) else [value]
                query = construct_query(key, terms)
                servers_to_update = servers_to_update.filter(query)

            total_updates = servers_to_update.count()
            if total_updates == 0:
                return JsonResponse({'status': 'warning', 'message': 'No servers to update with the given filters'}, status=200)

            bulk_comment = request.POST.get('bulk_comment', '')
            bulk_assigned_to = request.POST.get('bulk_assigned_to', '')

            def update_with_progress():
                batch_size = 50
                total_batches = (total_updates + batch_size - 1) // batch_size

                for i in range(total_batches):
                    batch_start = i * batch_size
                    batch_end = min((i + 1) * batch_size, total_updates)
                    batch = servers_to_update[batch_start:batch_end]

                    annotations = []
                    for server in batch:
                        annotation, created = DiscrepancyAnnotation.objects.get_or_create(SERVER_ID=server.SERVER_ID)
                        annotations.append(annotation)

                    for annotation in annotations:
                        annotation.add_entry(bulk_comment, bulk_assigned_to, request.user)

                    progress = ((i + 1) / total_batches) * 100
                    yield f"data: progress:{progress}|batch:{i + 1}\n\n"

            response = StreamingHttpResponse(update_with_progress(), content_type='text/event-stream')
            response['Cache-Control'] = 'no-cache'
            return response

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f"An unexpected error occurred: {e}"}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid Request Method'}, status=400)
    

@login_required
def chart_view(request):
    # Generate charts based on the filtered data
    
    json_data = get_field_labels()
    
    selected_fields = request.GET.getlist('fields')
    chart_types = request.GET.getlist('types')
    permanent_filter_selection = request.GET.get('permanentfilter')
    
    requestfilters = {}
    for key, value in request.GET.items():
        requestfilters[key] = value
    
    servers = get_filtered_servers(requestfilters, permanent_filter_selection)

    if "ANNOTATION" in selected_fields:
        server_ids = list(servers.values_list('SERVER_ID', flat=True).distinct())
        annotations = DiscrepancyAnnotation.objects.filter(SERVER_ID__in=server_ids)
        annotation_map = {ann.SERVER_ID: ann.comment for ann in annotations}
        
        fields_to_extract = ['SERVER_ID'] + [f for f in selected_fields if f != 'ANNOTATION']
        server_data = list(servers.values(*fields_to_extract).distinct())
        
        for server in server_data:
            server['ANNOTATION'] = annotation_map.get(server['SERVER_ID'], '')
    else:
        fields_to_extract = ['SERVER_ID'] + selected_fields
        server_data = list(servers.values(*fields_to_extract).distinct())

    # Pre-calculate the fields total here
    field_totals = {}
    for field in selected_fields:
       
        # Count the unique combos (SERVER_ID, valeur)
        unique_combos = set()
        for server in server_data:
            server_id = server.get('SERVER_ID')
            value = server.get(field, 'Unknown')
            if value is None or value == '':
                value = 'Unknown'
            unique_combos.add((server_id, str(value)))
        
        field_totals[field] = len(unique_combos)
    
    return generate_charts(request, server_data, json_data, selected_fields, chart_types, field_totals)   
    

def get_filtered_servers(requestfilters, permanent_filter_selection):
    # Filters servers based on the provided criteria and applies a permanent filter if selected
    
    json_path = os.path.join(os.path.dirname(__file__), 'field_labels.json')
    with open(json_path, "r", encoding="utf-8") as f:
        field_labels = json.load(f)

    # Initialize filters dictionary
    filters = {}

    # Populate filters dictionary for the fields
    for field_name, field_properties in field_labels['fields'].items():
        input_name = field_properties.get('inputname')
        if input_name:
            filters[field_name] = requestfilters.get(input_name, None)

    # Get all servers and apply filters
    servers = ServerDiscrepancy.objects.all().order_by('SERVER_ID')
    # Remove all [''] added during the filters creation
    filters = { k: v for k, v in filters.items() if v not in ['', None] }

    for key, value in filters.items():
        if isinstance(value, str) and ',' in value:
            terms = value.split(',')
        else:
            terms = [value]
        query = construct_query(key, terms)        
        servers = servers.filter(query)
    
    # Apply the permanent filter, if selected
    json_path=os.path.join(os.path.dirname(__file__), 'field_labels.json')
    with open(json_path, "r", encoding="utf-8") as f:
        json_data=json.load(f)        

    #permanent_filter_query, permanent_filter_names, permanent_filter_attributes = create_permanent_filter_query(json_data, permanent_filter_selection)
    #if permanent_filter_query:
    #    servers = servers.filter(permanent_filter_query)

    excluded_names = set(ExcludedServer.objects.values_list('server_name', flat=True))
    if excluded_names:
        servers = servers.exclude(SERVER_ID__in=excluded_names)

    return servers


def _get_filtered_servers_for_export(requestfilters):
    # Build a filtered ServerDiscrepancy queryset from URL params dict (inputname → value)
    json_data = get_field_labels()
    all_servers = ServerDiscrepancy.objects.all()
    combined_filter_query = Q()

    for field_key, field_info in json_data['fields'].items():
        if field_key in ('days_open', 'ANNOTATION'):
            continue
        input_name = field_info.get('inputname')
        if not input_name:
            continue
        raw = requestfilters.get(input_name, '')
        values = [v for v in (raw.split(',') if isinstance(raw, str) else raw) if v]
        if values:
            combined_filter_query &= construct_query(field_key, values)
            print(field_key)

    if combined_filter_query:
        all_servers = all_servers.filter(combined_filter_query)
        
    excluded_names = set(ExcludedServer.objects.values_list('server_name', flat=True))
    if excluded_names:
        all_servers = all_servers.exclude(SERVER_ID__in=excluded_names)
                

    # Apply sort order from URL params (same logic as server_view)
    sort_field = requestfilters.get('sort', 'SERVER_ID')
    sort_order = requestfilters.get('order', 'asc')

    if sort_field == 'days_open':
        oldest = DiscrepancyTracking.objects.filter(
            SERVER_ID=OuterRef('SERVER_ID')
        ).values('oldest_first_seen')[:1]
        all_servers = all_servers.annotate(oldest_issue=Subquery(oldest))
        order_expr = F('oldest_issue').asc(nulls_last=True) if sort_order == 'asc' else F('oldest_issue').desc(nulls_last=True)
    elif sort_field == 'ANNOTATION':
        order_expr = F('SERVER_ID').asc(nulls_last=True)
    else:
        valid_sort_fields = {k for k in json_data.get('fields', {}) if k not in ('days_open', 'ANNOTATION')}
        if sort_field not in valid_sort_fields:
            sort_field = 'SERVER_ID'
        order_expr = F(sort_field).asc(nulls_last=True) if sort_order == 'asc' else F(sort_field).desc(nulls_last=True)

    return all_servers.order_by(order_expr)
    

# Handles the export of server data to a specified file format (CSV or XLSX)
@login_required
def export_to_file(request, filetype):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not authorized'}, status=405)

    filetype = filetype.lower()
    if filetype not in ['xlsx', 'csv']:
        return JsonResponse({'error': 'Invalid format type'}, status=400)

    data = json.loads(request.body)
    requestfilters = data.get('filters', {})
    columns = data.get('columns', [])
    print(requestfilters)

    servers = _get_filtered_servers_for_export(requestfilters)
    hostname_list = list(servers.values_list('SERVER_ID', flat=True))

    # Annotations
    annotations_dict = {}
    if hostname_list:
        annotations = DiscrepancyAnnotation.objects.filter(SERVER_ID__in=hostname_list)
        annotations_dict = {ann.SERVER_ID: ann for ann in annotations}

    # Days open via _compute_days_open (handles oldest_first_seen + fallback)
    days_open_dict = {}
    if hostname_list:
        for t in DiscrepancyTracking.objects.filter(SERVER_ID__in=hostname_list):
            val = _compute_days_open(t)
            if val != '':
                days_open_dict[t.SERVER_ID] = val

    os.makedirs(EXPORT_DIR, exist_ok=True)
    job_id = str(uuid.uuid4())
    filepath = os.path.join(EXPORT_DIR, f"{job_id}.{filetype}")

    def background_export():
        try:
            if filetype == 'xlsx':
                generate_excel(filepath, servers, annotations_dict, days_open_dict, columns)
            else:
                generate_csv(filepath, servers, annotations_dict, days_open_dict, columns)
        except Exception as e:
            print(f"Error during export: {e}")

    threading.Thread(target=background_export).start()
    return JsonResponse({'job_id': job_id})


# Check the state of the file to export 
def export_status(request, job_id, filetype):

    extension = filetype.lower()
    if extension not in ['xlsx', 'csv']:
        return JsonResponse({'Error': 'Invalid filetype'}, status=400)

    # Create the full file name based on the job_id and extension passed in the URL from the s function startExport
    filename=f'{job_id}.{extension}'
    filepath = os.path.join(EXPORT_DIR, filename)
    
    # Return "ready" id the file is present, and "pending" if not. During its creation, the file as an additional .tmp extension to not trigger a "ready" too early
    if os.path.exists(filepath):
        return JsonResponse({'status': 'ready'})
    return JsonResponse({'status': 'pending'})


# Trigger the automatic downlaod when the export file is fully created
def download_export(request, job_id, filetype):

    from django.http import Http404
    extension = filetype.lower()
    if extension not in ['xlsx', 'csv']:
        return JsonResponse({'Error': 'Invalid filetype'}, status=400)
        
    # Create the full file name based on the job_id and extension passed in the URL from the s function startExport
    filename=f'{job_id}.{extension}'
    filepath = os.path.join(EXPORT_DIR, filename)
    
    # Raise an error if the file doesn't exist
    if not os.path.exists(filepath):
        raise Http404("File not found")

    # Initiate the automatic download
    response = FileResponse(open(filepath, 'rb'), as_attachment=True, filename=f'export_{app_name}.{extension}')

    def delete_file_after_close(resp):
        try:
            os.remove(filepath)
        except Exception as e:
            print(f"Error during the removing of the file : {e}")

    # Close and delete the generated file once downloaded
    response.close = lambda *args, **kwargs: delete_file_after_close(response)
    return response
        

@login_required
def save_search(request):
    if request.method == 'POST':
        search_name = request.POST.get('search_name', '').strip()
        if len(search_name) > 25:
            return JsonResponse({'success': False, 'message': 'Search name must be 25 characters or fewer.'})
        profile = UserProfile.objects.get(user=request.user)
        if SavedSearch.objects.filter(user_profile=profile, name=search_name).exists():
            return JsonResponse({'success': False, 'message': 'A search with this name already exists.'})
        filters = request.POST.get('filters')
        if not filters:
            return JsonResponse({'success': False, 'message': 'Filters cannot be empty.'})
        try:
            search_params = json.loads(filters)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid filter data.'})
        tags = json.loads(request.POST.get('tags', '[]'))
        SavedSearch.objects.create(
            user_profile=profile, name=search_name,
            filters=search_params, tags=tags, view=app_name
        )
        return redirect(request.META.get('HTTP_REFERER', '/'))
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def load_search(request, search_id):
    saved_search = SavedSearch.objects.get(id=search_id, user_profile__user=request.user)
    query_string = urlencode(saved_search.filters, doseq=True)
    return redirect(f'/{app_name}/servers/?{query_string}')


@login_required
def delete_search(request, search_id):
    saved_search = SavedSearch.objects.get(id=search_id, user_profile__user=request.user)
    saved_search.delete()
    return redirect(request.META.get('HTTP_REFERER', '/'))


# Display the logs_imports.html with the last 100 logs entries
def log_imports(request):

    # Get and sort the last entries of the table businesscontinuity_importstatus
    logs = ImportStatus.objects.order_by('-date_import')[:100]

    return render(request, f"{app_name}/logs_imports.html", {
        "logs": logs
    })


def exclusion_list_api(request):
    """
    GET  – return all excluded servers as JSON, with stale detection.
    POST – create entries (comma‑separated); validates against inventory.
    """

    if request.method == 'GET':
        from inventory.models import Server as InventoryServer

        exclusions = list(ExcludedServer.objects.all())
        excluded_names = [e.server_name for e in exclusions]
        upper_excluded = [name.upper() for name in excluded_names]

        inventory_qs = (
            InventoryServer.objects
            .annotate(_id_up=Upper('SERVER_ID'))
            .filter(_id_up__in=upper_excluded)
            .only('SERVER_ID', 'LIVE_STATUS', 'SNOW_STATUS')
        )
        server_map = {s.SERVER_ID.upper(): s for s in inventory_qs}
        data = []

        for e in exclusions:
            srv = server_map.get(e.server_name.upper())
            if srv is None:
                stale, stale_reason = True, 'not_found'
            elif str(srv.SNOW_STATUS or '').upper() in ('RETIRED', 'NON-OPERATIONAL') \
                    or str(srv.LIVE_STATUS or '').upper() not in ('ALIVE', 'LIVE'):
                stale, stale_reason = True, 'retired'
            else:
                stale, stale_reason = False, ''

            data.append({
                'id'            : e.id,
                'server_name'   : e.server_name,
                'reason'        : e.reason,
                'owner'         : e.owner,
                'exclusion_date': e.exclusion_date.strftime('%Y-%m-%d') if e.exclusion_date else '',
                'created_by'    : e.created_by,
                'created_at'    : e.created_at.strftime('%Y-%m-%d'),
                'stale'         : stale,
                'stale_reason'  : stale_reason,
            })

        return JsonResponse({'exclusions': data, 'count': len(data)})


    if request.method == 'POST':
        if not request.user.is_authenticated:
            return JsonResponse({'success': False,
                                 'message': 'Authentication required'}, status=401)

        server_names_raw = request.POST.get('server_names', '').strip()
        if not server_names_raw:
            return JsonResponse({'success': False,
                                 'message': 'Server name is required'}, status=400)

        server_names = [s.strip() for s in server_names_raw.split(',') if s.strip()]

        # This also gives us the list we will pass to the DB (upper‑cased)
        name_lookup = {s.upper(): s for s in server_names}
        upper_names = list(name_lookup.keys())          # ['SRV001', 'SRV002', 'FRAS000CFT96']

        from inventory.models import Server as InventoryServer

        existing_qs = (
            InventoryServer.objects
            .annotate(_id_up=Upper('SERVER_ID'))       # temporary column
            .filter(_id_up__in=upper_names)            # case‑insensitive IN
            .values_list('SERVER_ID', flat=True)      # returns the canonical DB value
        )

        existing_upper = {sid.upper() for sid in existing_qs}

        valid_names = [name_lookup[u] for u in upper_names if u in existing_upper]
        not_found   = [name_lookup[u] for u in upper_names if u not in existing_upper]

        if not valid_names:
            return JsonResponse({
                'success'          : False,
                'created'          : 0,
                'not_found'        : not_found,
                'already_excluded' : [],
                'message'          : 'None of the provided servers were found in inventory.',
            }, status=400)

        # Detect which valid names are already in the exclusion list
        already_excl_qs  = ExcludedServer.objects.filter(server_name__in=[n.upper() for n in valid_names])
        already_excl_map = {e.server_name.upper(): e for e in already_excl_qs}
        already_excluded_info = [
            {'id': already_excl_map[n.upper()].id, 'server_name': n,
             'reason': already_excl_map[n.upper()].reason, 'owner': already_excl_map[n.upper()].owner}
            for n in valid_names if n.upper() in already_excl_map
        ]
        to_create = [n for n in valid_names if n.upper() not in already_excl_map]

        reason = request.POST.get('reason', '').strip()
        owner  = request.POST.get('owner', '').strip()
        exclusion_date_str = request.POST.get('exclusion_date', '').strip()
        exclusion_date = parse_date(exclusion_date_str) if exclusion_date_str else None
        created_by = request.user.username

        for name in to_create:
            ExcludedServer.objects.create(
                server_name=name.upper(),
                reason=reason,
                owner=owner,
                exclusion_date=exclusion_date,
                created_by=created_by,
            )

        return JsonResponse({
            'success'          : len(to_create) > 0,
            'created'          : len(to_create),
            'not_found'        : not_found,
            'already_excluded' : already_excluded_info,
        })

    # Anything else → 405 Method Not Allowed
    return JsonResponse({'success': False,
                         'message': 'Invalid method'}, status=405)


def exclusion_delete_api(request, pk):
    """DELETE a single excluded server by primary key."""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    try:
        ExcludedServer.objects.get(pk=pk).delete()
        return JsonResponse({'success': True})
    except ExcludedServer.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)


def exclusion_update_api(request, pk):
    """Update reason/owner/date of an existing excluded server."""
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Authentication required'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST required'}, status=405)
    try:
        excl = ExcludedServer.objects.get(pk=pk)
    except ExcludedServer.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Not found'}, status=404)

    excl.reason = request.POST.get('reason', '').strip()
    excl.owner  = request.POST.get('owner', '').strip()
    exclusion_date_str = request.POST.get('exclusion_date', '').strip()
    excl.exclusion_date = parse_date(exclusion_date_str) if exclusion_date_str else None
    excl.save()
    return JsonResponse({'success': True})


def exclusion_export_csv(request):
    """Stream the full exclusions list as a CSV file."""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="exclusions.csv"'
    writer = csv.writer(response)
    writer.writerow(['Server Name', 'Reason', 'Owner', 'Excluded Until', 'Created By', 'Created On'])
    for e in ExcludedServer.objects.all():
        writer.writerow([
            e.server_name,
            e.reason,
            e.owner,
            e.exclusion_date.strftime('%Y-%m-%d') if e.exclusion_date else '',
            e.created_by,
            e.created_at.strftime('%Y-%m-%d, %H:%M'),
        ])
    return response    
  

def exclusion_export_excel(request):
    """Export the full exclusions list as an Excel file with stale/expired highlights."""
    import io
    import datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from inventory.models import Server as InventoryServer

    exclusions = list(ExcludedServer.objects.all())
    excluded_names = [e.server_name for e in exclusions]

    server_map = {}
    if excluded_names:
        server_map = {
            s.SERVER_ID: s
            for s in InventoryServer.objects.filter(SERVER_ID__in=excluded_names)
                                            .only('SERVER_ID', 'LIVE_STATUS', 'SNOW_STATUS')
        }

    today = dt.date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "Excluded Servers"
    ws.freeze_panes = 'A2'

    headers = ['Server Name', 'Reason', 'Owner', 'Excluded Until', 'Created By', 'Created On', 'Status']
    ws.append(headers)

    header_fill = PatternFill(start_color='A0826D', end_color='A0826D', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='D0C0B0'))
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[1].height = 20

    fill_expired = PatternFill(start_color='FFE5C8', end_color='FFE5C8', fill_type='solid')
    fill_stale   = PatternFill(start_color='FDE8E8', end_color='FDE8E8', fill_type='solid')

    for e in exclusions:
        srv = server_map.get(e.server_name)
        if srv is None:
            status, stale = 'Not in inventory', True
        elif str(srv.SNOW_STATUS or '').upper() in ('RETIRED', 'NON-OPERATIONAL') \
                or str(srv.LIVE_STATUS or '').upper() not in ('ALIVE', 'LIVE'):
            status, stale = 'Retired / Non-operational', True
        else:
            status, stale = 'OK', False

        expired = bool(e.exclusion_date and e.exclusion_date < today)

        ws.append([
            e.server_name,
            e.reason,
            e.owner,
            e.exclusion_date.strftime('%Y-%m-%d') if e.exclusion_date else '',
            e.created_by,
            e.created_at.strftime('%Y-%m-%d, %H:%M'),
            status,
        ])

        if stale or expired:
            row_fill = fill_stale if stale else fill_expired
            for cell in ws[ws.max_row]:
                cell.fill = row_fill

    col_widths = [22, 35, 16, 16, 16, 20, 24]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="exclusions.xlsx"'
    return response


@login_required
def dashboard_export_excel(request):
    """Export dashboard gauge data as Excel — one row per gauge."""
    import io
    import datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Reuse the same live-computation logic as dashboard_view ─────────────
    config_path = os.path.join(os.path.dirname(__file__), 'discrepancies_dashboard.json')
    with open(config_path, 'r') as f:
        config = json.load(f)

    from inventory.models import Server as InventoryServer

    pf_name = request.GET.get('pf', '').strip()
    json_data = get_field_labels()
    permanent_filters = json_data.get('permanentfilters', {})
    fields_info = json_data.get('fields', {})
    inputname_to_field = {finfo.get('inputname', fname.lower()): fname
                          for fname, finfo in fields_info.items()}

    excluded_names = set(ExcludedServer.objects.values_list('server_name', flat=True))

    _DASHBOARD_METRIC_FIELDS = {'alive_status_inconsistent', 'dead_status_inconsistent', 'missing_fields'}

    # Resolve filter_def — system permanent filter or ss: saved search
    filter_def = {}
    if pf_name and pf_name in permanent_filters:
        filter_def = permanent_filters[pf_name]
    elif pf_name and pf_name.startswith('ss:'):
        try:
            ss_id = int(pf_name[3:])
            ss = SavedSearch.objects.get(id=ss_id, user_profile__user=request.user)
            for inputname, value in ss.filters.items():
                field = inputname_to_field.get(inputname)
                if field and field not in _DASHBOARD_METRIC_FIELDS:
                    filter_def[field] = (
                        [v.strip() for v in value.split(',') if v.strip()]
                        if isinstance(value, str) else value
                    )
        except (ValueError, SavedSearch.DoesNotExist):
            pass
    elif pf_name and pf_name.startswith('multi:'):
        for part in pf_name[6:].split('|'):
            if ':' not in part:
                continue
            inputname, raw_values = part.split(':', 1)
            field = inputname_to_field.get(inputname)
            if field and field not in _DASHBOARD_METRIC_FIELDS:
                values = [v.strip() for v in raw_values.split(',') if v.strip()]
                if values:
                    filter_def[field] = values

    inv_field_names = {f.name for f in InventoryServer._meta.get_fields()}

    # Permanent filter Q
    inv_q = Q()
    disc_q = Q()
    for field, values in filter_def.items():
        disc_q &= construct_query(field, values)
        if field in inv_field_names:
            inv_q &= construct_query(field, values)

    is_filtered = bool(filter_def)
    disc_base = ServerDiscrepancy.objects.filter(disc_q) if is_filtered else ServerDiscrepancy.objects.all()
    if excluded_names:
        disc_base = disc_base.exclude(SERVER_ID__in=excluded_names)

    base_eligible_q = Q(LIVE_STATUS='ALIVE', SNOW_STATUS='OPERATIONAL', INFRAVERSION__in=['IV1', 'IV2', 'IBM'])
    inv_eligible = InventoryServer.objects.filter(base_eligible_q & inv_q) if is_filtered else InventoryServer.objects.filter(base_eligible_q)
    if excluded_names:
        inv_eligible = inv_eligible.exclude(SERVER_ID__in=excluded_names)
    total_eligible = inv_eligible.values('SERVER_ID').distinct().count()
    total_physical = inv_eligible.filter(MACHINE_TYPE='PHYSICAL').values('SERVER_ID').distinct().count()

    inv_all = InventoryServer.objects.filter(inv_q) if is_filtered else InventoryServer.objects.all()
    if excluded_names:
        inv_all = inv_all.exclude(SERVER_ID__in=excluded_names)
    total_all = inv_all.values('SERVER_ID').distinct().count()

    METRIC_TO_FIELD = {
        'missing_live_status_count':       'LIVE_STATUS',
        'missing_osshortname_count':       'OSSHORTNAME',
        'missing_osfamily_count':          'OSFAMILY',
        'missing_snow_supportgroup_count': 'SNOW_SUPPORTGROUP',
        'missing_machine_type_count':      'MACHINE_TYPE',
        'missing_manufacturer_count':      'MANUFACTURER',
        'missing_country_count':           'COUNTRY',
        'missing_app_auid_value_count':    'APP_AUID_VALUE',
        'missing_app_name_value_count':    'APP_NAME_VALUE',
        'missing_region_count':            'REGION',
        'missing_city_count':              'CITY',
        'missing_infraversion_count':      'INFRAVERSION',
        'missing_ipaddress_count':         'IPADDRESS',
        'missing_snow_status_count':       'SNOW_STATUS',
        'missing_idrac_name_count':        'IDRAC_NAME',
        'missing_idrac_ip_count':          'IDRAC_IP',
    }

    # Build rows
    rows = []
    for widget in config['dashboard']['widgets']:
        wid   = widget['id']
        title = widget['title'].replace('\n', ' ')
        metric = widget.get('metric', '')
        physical_only = widget.get('physical_only', 'no') == 'yes'

        if widget['size'] == 'large':
            if wid == 'missing_data':
                issues = disc_base.filter(missing_fields__gt='').count()
                total  = total_eligible
            else:  # operational_inconsistencies
                issues = disc_base.filter(
                    Q(alive_status_inconsistent='KO') | Q(dead_status_inconsistent='KO')
                ).count()
                total = total_all
        else:
            if metric == 'alive_status_inconsistent_count':
                issues = disc_base.filter(alive_status_inconsistent='KO').count()
            elif metric == 'dead_status_inconsistent_count':
                issues = disc_base.filter(dead_status_inconsistent='KO').count()
            else:
                field = METRIC_TO_FIELD.get(metric)
                issues = disc_base.filter(missing_fields__icontains=field).count() if field else 0
            total = total_physical if physical_only else total_eligible

        ok  = max(0, total - issues)
        pct = math.trunc((ok / total * 10000)) / 100 if total else 100
        rows.append((widget['size'], title, pct, ok, issues, total))

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.freeze_panes = 'A3'

    # Title row
    if pf_name and pf_name in permanent_filters:
        filter_label = pf_name
    elif pf_name and pf_name.startswith('ss:') and filter_def:
        try:
            ss_id = int(pf_name[3:])
            ss_obj = SavedSearch.objects.get(id=ss_id, user_profile__user=request.user)
            filter_label = ss_obj.name
        except (ValueError, SavedSearch.DoesNotExist):
            filter_label = 'All Servers'
    elif pf_name and pf_name.startswith('multi:') and filter_def:
        label_parts = []
        for part in pf_name[6:].split('|'):
            if ':' in part:
                label_parts.append(part.split(':', 1)[1].replace(',', ', '))
        filter_label = ' + '.join(label_parts) if label_parts else 'Multi-filter'
    else:
        filter_label = 'All Servers'
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Server Data Quality Dashboard — {filter_label}  |  {dt.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    title_cell.font = Font(bold=True, size=11, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='A0826D', end_color='A0826D', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 22

    # Header row
    headers = ['Title', '% OK', 'OK', 'Issues', 'Total']
    ws.append(headers)
    hdr_fill = PatternFill(start_color='D6C4BB', end_color='D6C4BB', fill_type='solid')
    for cell in ws[2]:
        cell.font = Font(bold=True, size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    # Color fills based on thresholds (critical<98, warning<99, good>=99)
    fill_good     = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    fill_warning  = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    fill_critical = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    fill_large    = PatternFill(start_color='EDE5DF', end_color='EDE5DF', fill_type='solid')

    thin = Side(style='thin', color='D0C0B0')

    prev_size = None
    for size, title, pct, ok, issues, total in rows:
        # Blank separator row between large and small gauges
        if prev_size == 'large' and size == 'small':
            ws.append([''])
            ws.row_dimensions[ws.max_row].height = 8

        ws.append([title, pct / 100, ok, issues, total])
        row = ws[ws.max_row]

        # Row background
        if size == 'large':
            bg = fill_large
        elif pct >= 99:
            bg = fill_good
        elif pct >= 98:
            bg = fill_warning
        else:
            bg = fill_critical

        for cell in row:
            cell.fill = bg
            cell.alignment = Alignment(vertical='center')
            cell.border = Border(bottom=thin)

        # Bold + larger font for large gauges
        font_size = 11 if size == 'large' else 10
        for cell in row:
            cell.font = Font(bold=(size == 'large'), size=font_size)

        # % column formatted as percentage
        row[1].number_format = '0.00%'
        row[1].alignment = Alignment(horizontal='center', vertical='center')
        for i in (2, 3, 4):
            row[i].alignment = Alignment(horizontal='right', vertical='center')

        ws.row_dimensions[ws.max_row].height = 18 if size == 'large' else 16
        prev_size = size

    # Column widths
    for col, width in zip('ABCDE', [36, 10, 10, 10, 10]):
        ws.column_dimensions[col].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    slug = filter_label.lower().replace(' ', '_')
    filename = f"dashboard_{slug}_{dt.datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
