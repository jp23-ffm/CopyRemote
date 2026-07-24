import csv
import itertools
import os

from django.conf import settings
from openpyxl import Workbook

EXPORT_DIR = os.path.join(settings.MEDIA_ROOT, 'exports')

# Rows pulled from the DB per round-trip when streaming a queryset.
CHUNK_SIZE = 2000


def get_db_attr(obj, attr):
    parts = attr.split('.')
    for part in parts:
        if hasattr(obj, part):
            obj = getattr(obj, part)
        else:
            return None
    return obj


def clean_value(value):
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
    text = ' '.join(text.split())
    return text


def generate_excel(filepath, servers, annotations_dict, columns, FILTER_MAPPING, su_dict=None, su_fields=None):
    su_fields = su_fields or set()
    temp_filepath = filepath + '.tmp'
    # write_only streams rows straight to the zip on disk instead of keeping
    # every cell in memory until save() — needed for large exports.
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Servers")

    ws.append(columns)

    if servers.exists():
        for server in servers.iterator(chunk_size=CHUNK_SIZE):
            row = []
            for col in columns:
                if col == 'ANNOTATION':
                    row.append(clean_value(annotations_dict.get(server.SERVER_ID, '')))
                elif col in su_fields:
                    su = su_dict.get(server.SERVER_ID) if su_dict else None
                    row.append(getattr(su, col, '') if su else '')
                else:
                    row.append(get_db_attr(server, FILTER_MAPPING[col]))
            ws.append(row)
    else:
        print("No servers matching the applied filters were found.")

    wb.save(temp_filepath)
    os.replace(temp_filepath, filepath)


def generate_csv(filepath, servers, annotations_dict, columns, FILTER_MAPPING, su_dict=None, su_fields=None):
    su_fields = su_fields or set()
    temp_filepath = filepath + '.tmp'
    with open(temp_filepath, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)

        writer.writerow(columns)

        for server in servers.iterator(chunk_size=CHUNK_SIZE):
            row = []
            for col in columns:
                if col == 'ANNOTATION':
                    row.append(clean_value(annotations_dict.get(server.SERVER_ID, '')))
                elif col in su_fields:
                    su = su_dict.get(server.SERVER_ID) if su_dict else None
                    row.append(getattr(su, col, '') if su else '')
                else:
                    row.append(get_db_attr(server, FILTER_MAPPING[col]))
            writer.writerow(row)

    os.replace(temp_filepath, filepath)


def _iter_grouped(servers_ordered):
    """Stream (SERVER_ID, [rows]) groups. servers_ordered must be ordered by SERVER_ID
    so that all rows of a given server are contiguous and only one group at a time
    needs to be held in memory."""
    for hostname, group in itertools.groupby(
        servers_ordered.iterator(chunk_size=CHUNK_SIZE), key=lambda s: s.SERVER_ID
    ):
        yield hostname, list(group)


def _build_grouped_row(hostname, server_list, summary, columns, annotations_dict, su_dict, su_fields):
    row = []

    if len(server_list) == 1:
        single_server = server_list[0]
        for field_name in columns:
            if field_name == 'ANNOTATION':
                row.append(clean_value(annotations_dict.get(hostname, '')))
            elif field_name in su_fields:
                su = su_dict.get(hostname) if su_dict else None
                row.append(clean_value(getattr(su, field_name, '') if su else ''))
            else:
                row.append(clean_value(getattr(single_server, field_name, '')))
    else:
        if summary:
            for field_name in columns:
                if field_name == 'ANNOTATION':
                    row.append(clean_value(annotations_dict.get(hostname, '')))
                elif field_name in su_fields:
                    su = su_dict.get(hostname) if su_dict else None
                    row.append(clean_value(getattr(su, field_name, '') if su else ''))
                elif field_name in summary.constant_fields:
                    row.append(clean_value(summary.constant_fields[field_name]))
                elif field_name in summary.variable_fields:
                    preview = summary.variable_fields[field_name].get('preview', '')
                    row.append(clean_value(preview))
                else:
                    row.append('')
        else:
            row.extend([''] * len(columns))

    return row


def generate_csv_grouped(filepath, servers_ordered, summaries_dict, annotations_dict, columns, su_dict=None, su_fields=None):
    su_fields = su_fields or set()
    temp_filepath = filepath + '.tmp'
    with open(temp_filepath, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)

        writer.writerow(columns)

        for hostname, server_list in _iter_grouped(servers_ordered):
            summary = summaries_dict.get(hostname)
            row = _build_grouped_row(hostname, server_list, summary, columns, annotations_dict, su_dict, su_fields)
            writer.writerow(row)

    os.replace(temp_filepath, filepath)


def generate_excel_grouped(filepath, servers_ordered, summaries_dict, annotations_dict, columns, su_dict=None, su_fields=None):
    su_fields = su_fields or set()
    temp_filepath = filepath + '.tmp'
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Servers")

    ws.append(columns)

    for hostname, server_list in _iter_grouped(servers_ordered):
        summary = summaries_dict.get(hostname)
        row = _build_grouped_row(hostname, server_list, summary, columns, annotations_dict, su_dict, su_fields)
        ws.append(row)

    wb.save(temp_filepath)
    os.replace(temp_filepath, filepath)
