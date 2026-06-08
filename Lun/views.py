import csv
import datetime
import django
import io
import json
import os
import time
import threading
import uuid
import socket
from functools import lru_cache

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.cache import cache
from django.db import models, transaction, connection
from django.db.models import Q, F, Case, When, Count, OuterRef, Subquery, Min, Max
from django.http import FileResponse, JsonResponse, HttpResponse, HttpResponseRedirect, StreamingHttpResponse, QueryDict
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt

from accessrights.helpers import has_perm
from businesscontinuity.models import ServerUnique
from common.views import generate_charts
from userapp.models import UserProfile, SavedSearch, SavedOptions, UserPermissions, SavedChart, SavedHistoryQuery
from .exports import generate_csv, generate_csv_grouped, generate_excel, generate_excel_grouped, EXPORT_DIR
from .models import Server, ServerGroupSummary, ServerAnnotation, ImportStatus, FieldSnapshot, FieldSnapshotFiltered, SnapshotStatus, ServerHistory
from .forms import AnnotationForm, CsvUploadForm

from collections import defaultdict
from urllib.parse import urlencode, parse_qs, unquote
from threading import Lock


app_name = __package__.split('.')[-1]


# Global cache for field labels with thread-safe access
_field_labels_cache = None
_field_labels_file_mtime = 0
_cache_lock = Lock()


def get_field_labels():
    # Cache field_labels.json with update when the file changes, Thread-safe for Gunicorn

    global _field_labels_cache, _field_labels_file_mtime
    
    json_path = os.path.join(os.path.dirname(__file__), 'field_labels.json')
    
    try:
        current_file_mtime = os.path.getmtime(json_path)
    except OSError:
        return _field_labels_cache

    if _field_labels_cache is not None and current_file_mtime == _field_labels_file_mtime:
        return _field_labels_cache
    
    with _cache_lock:
        try:
            current_file_mtime = os.path.getmtime(json_path)
        except OSError:
            return _field_labels_cache
        
        if _field_labels_cache is not None and current_file_mtime == _field_labels_file_mtime:
            return _field_labels_cache
        
        try:
            with open(json_path, 'r', encoding="utf-8") as f:
                _field_labels_cache = json.load(f)
            
            _field_labels_file_mtime = current_file_mtime
            
        except (OSError, json.JSONDecodeError) as e:
            return _field_labels_cache
    
    return _field_labels_cache


def invalidate_field_labels_cache():
    # Call this when field_labels.json is modified
    global _field_labels_cache, _field_labels_file_mtime
    with _cache_lock:
        _field_labels_cache = None
        _field_labels_file_mtime = 0


# Listbox caching

def get_all_listbox_fields(json_data):
    # Extract all fields that need listbox values
    listbox_fields = []
    for key, val in json_data.get('fields', {}).items():
        if val.get('listbox'):
            listbox_fields.append(key)
    return listbox_fields


def get_serverunique_fields(json_data=None):
    if json_data is None:
        json_data = get_field_labels()
    return {k for k, v in json_data.get('fields', {}).items() if v.get('model_extra') == 'serverunique'}


def batch_load_listbox_values(listbox_fields, su_fields=None, force_refresh=False):
    # Load all listbox values in a single optimized query batch

    su_fields = su_fields or set()
    cache_key_prefix = 'listbox_'
    result = {}
    fields_to_query = []

    # Check cache first
    for field in listbox_fields:
        cache_key = f'{cache_key_prefix}{field}'
        if not force_refresh:
            cached = cache.get(cache_key)
            if cached is not None:
                result[field] = cached
                continue
        fields_to_query.append(field)

    if not fields_to_query:
        return result

    # Batch query for all uncached fields using raw SQL for efficiency
    with connection.cursor() as cursor:
        for field in fields_to_query:
            # Use the correct table based on whether this is a BC field
            table = 'businesscontinuity_serverunique' if field in su_fields else 'inventory_server'
            cursor.execute(f'''
                SELECT DISTINCT "{field}"
                FROM {table}
                WHERE "{field}" IS NOT NULL AND "{field}" != ''
                ORDER BY "{field}"
            ''')
            values = [row[0] for row in cursor.fetchall()]

            # Sort with EMPTY at end
            if any(v and v.upper() == "EMPTY" for v in values):
                values = [v for v in values if v and v.upper() != "EMPTY"]
                values.sort()
                values.append("EMPTY")

            result[field] = values

            # Cache for 1 hour
            cache.set(f'{cache_key_prefix}{field}', values, 3600)

    return result


def construct_query(key, terms):
    # Creates a Django Q object based on a list of terms for a specific field
    query = Q()
    
    # Iterate over each term in the terms list
    for term in terms:
        if term.startswith('@'):  # Check if the term starts with '@' for an exact match
            term = term[1:]  # Remove the '@' character
            query |= Q(**{f'{key}__iexact': term})  # Create a Q object for an exact match (case-insensitive)
        elif term.startswith('!'):  # Check if the term starts with '!' for an exclusion
            term = term[1:]  # Remove the '!' character
            query &= ~Q(**{f'{key}__icontains': term})  # Create a Q object for an exclusion (case-insensitive containment test)
        else:  # For terms without special characters, perform a containment test
            query |= Q(**{f'{key}__icontains': term})  # Create a Q object for a containment test (case-insensitive)

    # Return the combined Q object representing the filter criteria
    return query


def create_permanent_filter_query(json_data, selected_option):
    # Create a permanent filter query based on predefined filters in JSON configuration

    permanent_filters = json_data.get("permanentfilters", {})
    permanent_filter_attributes = permanent_filters.get(selected_option) if selected_option else None

    overall_query = Q()

    if permanent_filter_attributes:
        for key, value in permanent_filter_attributes.items():
            terms = [value] if isinstance(value, str) else value
            query = construct_query(key, terms)
            overall_query &= query

    return overall_query, permanent_filters, permanent_filter_attributes


def build_filters_from_request(request, json_data):
    # Build filters dictionary from request parameters

    filters = {}

    for field_key, field_info in json_data['fields'].items():
        input_name = field_info.get('inputname')
        if not input_name:
            continue
        if field_info.get('fieldtype') == 'date':
            date_from = request.GET.get(f'{input_name}_from', '').strip()
            date_to = request.GET.get(f'{input_name}_to', '').strip()
            if date_from:
                filters[f'{field_key}__gte'] = [date_from]
            if date_to:
                filters[f'{field_key}__lte'] = [date_to]
        else:
            filter_value = request.GET.get(input_name, '').split(',')
            filters[field_key] = [v for v in filter_value if v]                

    return filters


def apply_filters_to_queryset(queryset, filters, su_fields=None):
    # Apply filters to a queryset

    su_fields = su_fields or set()
    combined_filter_query = Q()
    su_filter_query = Q()

    for key, values in filters.items():
        if not values or key == "ANNOTATION":
            continue

        if key.endswith('__gte') or key.endswith('__lte'):
            field_name, lookup = key.rsplit('__', 1)
            combined_filter_query &= Q(**{f'{field_name}__{lookup}': values[0]})
        else:
            query = construct_query(key, values)
            if key in su_fields:
                su_filter_query &= query
            else:
                combined_filter_query &= query


    if combined_filter_query:
        queryset = queryset.filter(combined_filter_query)

    if su_filter_query:
        matching_hostnames = ServerUnique.objects.filter(su_filter_query).values_list('hostname', flat=True)
        queryset = queryset.filter(SERVER_ID__in=matching_hostnames)

    return queryset


#  MAIN VIEW

@login_required
def server_view(request):
    
    start_time = time.time()
    localhostname = socket.gethostname()

    # Get user profile (create if needed)
    profile, _ = UserProfile.objects.get_or_create(user=request.user)

    # Access rights
    can_continuity = has_perm(request.user, 'inventory.continuity')
    can_annotation = has_perm(request.user, 'inventory.annotations')

    # Check edit permissions
    try:
        user_permissions = UserPermissions.objects.get(user_profile=profile)
        edit_mode_user_permissions = user_permissions.inventory_allowedit
    except UserPermissions.DoesNotExist:
        edit_mode_user_permissions = False
    edit_mode = edit_mode_user_permissions or can_annotation

    # Load cached JSON configuration
    json_data = get_field_labels()

    # Get permanent filter selection
    json_data_permanentfilters = json_data.get("permanentfilters", {})
    try:
        user_options = SavedOptions.objects.get(user_profile=profile)
        permanent_filter_selection = user_options.inventory_permanentfilter
    except SavedOptions.DoesNotExist:
        permanent_filter_selection = "All Servers"

    if permanent_filter_selection != "All Servers" and permanent_filter_selection not in json_data_permanentfilters:
        permanent_filter_selection = "All Servers"

    # Build permanent filter query
    permanent_filter_query, permanent_filter_names, permanent_filter_attributes = \
        create_permanent_filter_query(json_data, permanent_filter_selection)

    # Build filters from request
    filters = build_filters_from_request(request, json_data)

    # Get visible columns from request (for lazy loading optimization)
    visible_columns_param = request.GET.get('visible_columns', '')
    visible_columns = set(visible_columns_param.split(',')) if visible_columns_param else None

    # Base queryset - only order once
    base_queryset = Server.objects.all()

    # Apply permanent filter
    if permanent_filter_query:
        base_queryset = base_queryset.filter(permanent_filter_query)

    # Apply user filters
    su_fields = get_serverunique_fields(json_data)
    base_queryset = apply_filters_to_queryset(base_queryset, filters, su_fields=su_fields)

    # Handle annotation filter specially
    annotation_filter = request.GET.get("annotation")
    if annotation_filter:
        annotation_terms = annotation_filter.split(',')
        query = construct_query("notes", annotation_terms)
        server_ids = ServerAnnotation.objects.filter(query).values_list('SERVER_ID', flat=True)
        base_queryset = base_queryset.filter(SERVER_ID__in=server_ids)

    # Sort handling
    sort_field = request.GET.get('sort', 'SERVER_ID')
    sort_order = request.GET.get('order', 'asc')
    unsortable = su_fields | {'ANNOTATION'}
    valid_sort_fields = {f_key for f_key in json_data.get('fields', {}).keys()} - unsortable
    if sort_field not in valid_sort_fields:
        sort_field = 'SERVER_ID'
    order_expr = F(sort_field).asc(nulls_last=True) if sort_order == 'asc' else F(sort_field).desc(nulls_last=True)

    # Final ordering: default = SERVER_ID + APP_NAME_VALUE, explicit sort = field + SERVER_ID only
    explicit_sort = 'sort' in request.GET
    if explicit_sort:
        filtered_servers = base_queryset.order_by(order_expr, 'SERVER_ID')
    else:
        filtered_servers = base_queryset.order_by('SERVER_ID', 'APP_NAME_VALUE')

    # Pagination settings
    page_size = int(request.GET.get('page_size', 50))
    page_number = request.GET.get("page")

    # Prepare model fields info
    json_data_fields = json_data.get("fields", {})
    if 'ANNOTATION' in json_data_fields and not edit_mode:
        json_data_fields = {k: v for k, v in json_data_fields.items() if k != 'ANNOTATION'}
    if not can_continuity:
        json_data_fields = {k: v for k, v in json_data_fields.items() if k not in su_fields}

    model_fields = []
    for field_name, field_info in json_data_fields.items():
        model_fields.append({
            'name': field_name,
            'verbose_name': field_info.get('displayname', field_name.replace('_', ' ').title()),
            'is_hostname': field_name == 'SERVER_ID'
        })

    #  Data fetching

    display_servers = []
    cacheset = False

    # Query 1: Get distinct hostnames for pagination
    # For explicit sort on a non-SERVER_ID field, aggregate per hostname
    if explicit_sort and sort_field != 'SERVER_ID':
        agg_func = Min if sort_order == 'asc' else Max
        sort_expr = F('sort_val').asc(nulls_last=True) if sort_order == 'asc' else F('sort_val').desc(nulls_last=True)
        filtered_hostnames_qs = (
            filtered_servers
            .values('SERVER_ID')
            .annotate(sort_val=agg_func(sort_field))
            .order_by(sort_expr, 'SERVER_ID')
        )
    else:
        filtered_hostnames_qs = (
            filtered_servers
            .values('SERVER_ID')
            .distinct()
            .order_by(order_expr)
        )

    hostnames_paginator = Paginator(filtered_hostnames_qs, page_size)
    hostnames_page = hostnames_paginator.get_page(page_number)
    hostnames_in_page = [item['SERVER_ID'] for item in hostnames_page]

    if hostnames_in_page:
        # Query 2: Get all data in ONE query - servers, summaries, and annotations
        servers_for_page = filtered_servers.filter(SERVER_ID__in=hostnames_in_page)
        servers_list = list(servers_for_page)

        # Query 3: Summaries (uses only() for efficiency)
        summaries_queryset = ServerGroupSummary.objects.filter(
            SERVER_ID__in=hostnames_in_page
        ).only('SERVER_ID', 'total_instances', 'constant_fields', 'variable_fields')
        summaries_dict = {s.SERVER_ID: s for s in summaries_queryset}

        # Query 4: Annotations (only if edit mode)
        annotations_dict = {}
        if edit_mode:
            annotations = ServerAnnotation.objects.filter(SERVER_ID__in=hostnames_in_page)
            annotations_dict = {ann.SERVER_ID: ann for ann in annotations}

        # Group servers by hostname
        server_groups = defaultdict(list)
        for server in servers_list:
            server_groups[server.SERVER_ID].append(server)

        # Build display data maintaining page order
        for SERVER_ID in hostnames_in_page:
            server_list = server_groups.get(SERVER_ID, [])
            if not server_list:
                continue

            summary = summaries_dict.get(SERVER_ID)

            if summary:
                visible_count = len(server_list)
                total_count = summary.total_instances
                hidden_count = max(0, total_count - visible_count)

                if visible_count == 1:
                    # Single instance - show real data
                    single_server = server_list[0]
                    constant_fields = {}
                    for field in single_server._meta.fields:
                        if field.name not in ['id', 'created_at', 'updated_at']:
                            value = getattr(single_server, field.name)
                            if value:
                                constant_fields[field.name] = str(value)

                    display_servers.append({
                        'hostname': SERVER_ID,
                        'count': visible_count,
                        'total_count': total_count,
                        'hidden_count': hidden_count,
                        'has_hidden': hidden_count > 0,
                        'constant_fields': constant_fields,
                        'variable_fields': {},
                        'all_instances': server_list,
                        'annotation': annotations_dict.get(SERVER_ID)
                    })
                else:
                    # Multiple instances - use summary
                    display_servers.append({
                        'hostname': SERVER_ID,
                        'count': visible_count,
                        'total_count': total_count,
                        'hidden_count': hidden_count,
                        'has_hidden': hidden_count > 0,
                        'constant_fields': summary.constant_fields,
                        'variable_fields': summary.variable_fields,
                        'all_instances': server_list,
                        'annotation': annotations_dict.get(SERVER_ID),
                        'instances_json': json.dumps([{
                            'constant_fields': {fn: str(getattr(s, fn, '')) for fn in summary.constant_fields.keys()},
                            'variable_fields': {fn: str(getattr(s, fn, '')) for fn in summary.variable_fields.keys()},
                        } for s in server_list], ensure_ascii=False),
                    })
            else:
                display_servers.append({
                    'hostname': SERVER_ID,
                    'count': len(server_list),
                    'total_count': len(server_list),
                    'hidden_count': 0,
                    'has_hidden': False,
                    'constant_fields': {'status': 'Summary missing - please rebuild'},
                    'variable_fields': {},
                    'all_instances': server_list,
                    'annotation': annotations_dict.get(SERVER_ID)
                })

    # Enrich with ServerUnique data for BC fields
    needs_su = visible_columns and any(f in visible_columns for f in su_fields)
    if needs_su and hostnames_in_page:
        su_dict = {su.hostname: su for su in ServerUnique.objects.filter(hostname__in=hostnames_in_page)}

        for server_group in display_servers:
            server_group['serverunique'] = su_dict.get(server_group['hostname'])

    page_obj = create_page_wrapper(display_servers, hostnames_page)

    # Statistics
    total_servers_stat = hostnames_paginator.count


    # Get all listbox fields
    listbox_fields = get_all_listbox_fields(json_data)

    # Batch load all listbox values
    listbox_values = batch_load_listbox_values(listbox_fields, su_fields=su_fields)
    cacheset = bool(listbox_values)

    # Build table_fields with listbox values
    table_fields = []
    for key, val in json_data_fields.items():
        listbox_value = val.get("listbox", '')
        if listbox_value and key in listbox_values:
            listbox_evaluated = listbox_values[key]
        else:
            listbox_evaluated = ''

        table_fields.append({
            "name": key,
            "displayname": val.get("displayname", key),
            "inputname": val.get("inputname", key),
            "listbox": listbox_evaluated,
            "listboxmsg": val.get("listboxmsg", 'Select an option'),
            "listid": val.get("listid", 'missingid'),
            "fieldtype": val.get("fieldtype", '')
        })

    # Build category fields
    json_data_categories = json_data.get("categories", {})
    grouped = defaultdict(list)
    for key, value in json_data_fields.items():
        if isinstance(value, dict):
            section = value.get('selectionsection', '').strip()
            displayname = value.get('displayname', '').strip() or key
            ischecked = value.get('ischecked') == "True"
            ischeckeddisabled = value.get('ischeckeddisabled') == "True"

            if section in json_data_categories:
                grouped[section].append({
                    'key': key,
                    'displayname': displayname,
                    'ischecked': ischecked,
                    'ischeckeddisabled': ischeckeddisabled
                })

    category_fields = [
        {
            'category': cat,
            'title': json_data_categories[cat],
            'fields': grouped[cat]
        }
        for cat in json_data_categories
        if cat in grouped
    ]

    # Other context data
    saved_searches = profile.savedsearch_set.filter(view=app_name)
    last_status = ImportStatus.objects.order_by('-date_import').first()

    fields_labels = json_data.get('fields', {})
    permanent_filters_tooltip = {}
    for fname, conditions in permanent_filter_names.items():
        parts = []
        for field, values in conditions.items():
            display = fields_labels.get(field, {}).get('displayname', field)
            parts.append(f"{display}: {', '.join(values)}")
        permanent_filters_tooltip[fname] = ' | '.join(parts)


    context = {
        'page_obj': page_obj,
        'table_fields': table_fields,
        'category_fields': category_fields,
        'permanent_filters_fields': permanent_filter_names,
        'permanent_filters_tooltip': permanent_filters_tooltip,        
        'appname': app_name,
        'page_size': page_size,
        'saved_searches': saved_searches,
        'last_status': last_status,
        'current_filters': filters,
        'json_data': json.dumps(json_data),
        'loggedonuser': request.user,
        'permanent_filter_selection': permanent_filter_selection,
        'model_fields': model_fields,
        'total_servers': total_servers_stat,

        'flat_view': False,
        'edit_mode': edit_mode,
        'cacheset': cacheset,
        'localhostname': localhostname,
        'visible_columns': visible_columns_param,
        'sort_field': sort_field,
        'sort_order': sort_order,
        'can_continuity': can_continuity,
    }

    return render(request, f'{app_name}/servers.html', context)


def create_page_wrapper(object_list, source_page):
    # Create a unified page object wrapper for both modes"%
    class UnifiedPageObj:
        def __init__(self, objects, source):
            self.object_list = objects
            self.number = source.number
            self.paginator = source.paginator

        def __iter__(self):
            return iter(self.object_list)

        def has_previous(self):
            return self.number > 1

        def has_next(self):
            return self.number < self.paginator.num_pages

        def previous_page_number(self):
            return self.number - 1 if self.has_previous() else None

        def next_page_number(self):
            return self.number + 1 if self.has_next() else None

        def has_other_pages(self):
            return self.paginator.num_pages > 1

    return UnifiedPageObj(object_list, source_page)


def _load_relation_config():
    config_path = os.path.join(os.path.dirname(__file__), 'relation_config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


@login_required
@require_http_methods(["GET"])
def server_relations(request, hostname):
    """
    Build a structured relation graph for a server.

    Rules:
    - type=server  : direct link from center (hypervisor); deduplicated
    - type=property, no parent:
        * single distinct value  -> direct link center -> value node
        * multiple distinct values -> group node center -> [Group] -> value nodes
    - type=property, parent=FIELD:
        * values attach to the corresponding parent-value node, determined
          per-instance (App1->DEV, App2->PROD, not App1/App2->DEV+PROD)
    """
    config = _load_relation_config()

    servers = Server.objects.filter(SERVER_ID=hostname)
    if not servers.exists():
        return JsonResponse({'error': 'Server not found'}, status=404)

    server_count = servers.count()

    # Load all instances as plain dicts once
    rel_fields = [r['field'] for r in config.get('relations', [])]
    instances  = list(servers.values(*rel_fields))

    # Pre-compute per-field: distinct values + {value -> set(instance indices)}
    field_meta = {}   # field -> {'values': [...], 'idx_map': {value: {0,1,…}}}
    for rel in config.get('relations', []):
        if rel['type'] != 'property':
            continue
        field   = rel['field']
        idx_map = {}
        for i, inst in enumerate(instances):
            v = (inst.get(field) or '').strip()
            if v:
                idx_map.setdefault(v, set()).add(i)
        field_meta[field] = {
            'values':  list(idx_map.keys()),
            'idx_map': idx_map,
        }

    nodes         = [{'id': hostname, 'label': hostname, 'type': 'center', 'count': server_count}]
    links         = []
    seen_node_ids = {hostname}
    seen_servers  = {}   # server value -> node_id  (dedup VPIC_HOST / HYPERVISOR)

    def add_node(n):
        if n['id'] not in seen_node_ids:
            nodes.append(n)
            seen_node_ids.add(n['id'])

    def add_link(src, tgt):
        links.append({'source': src, 'target': tgt})

    # Server-type relations (hypervisors)
    for rel in config.get('relations', []):
        if rel['type'] != 'server':
            continue
        field = rel['field']
        for inst in instances:
            value = (inst.get(field) or '').strip()
            if not value or value in seen_servers:
                continue
            pivot_count = Server.objects.filter(SERVER_ID=value).count()
            node_id = f'server__{value}'
            seen_servers[value] = node_id
            add_node({
                'id':          node_id,
                'label':       value,
                'field_label': rel['label'],
                'type':        'server',
                'color':       rel.get('color', 'hypervisor'),
                'pivot_id':    value if pivot_count > 0 else None,
                'count':       pivot_count,
            })
            add_link(hostname, node_id)

    # Property relations (all flat — direct link from center)
    # All property fields connect directly to the server node (star topology).
    # parent config is ignored in this view; each field gets its own direct edge.
    for rel in config.get('relations', []):
        if rel['type'] != 'property':
            continue
        field  = rel['field']
        meta   = field_meta.get(field, {})
        values = meta.get('values', [])
        if not values:
            continue

        if len(values) == 1:
            # Single value -> direct link to center
            v           = values[0]
            node_id     = f'prop__{field}__{v}'
            child_count = Server.objects.filter(**{field: v}).values('SERVER_ID').distinct().count()
            add_node({'id': node_id, 'label': v, 'field_label': rel['label'],
                      'type': 'property', 'color': rel.get('color', 'property'),
                      'field': field, 'value': v, 'child_count': child_count})
            add_link(hostname, node_id)
        else:
            # Multiple values -> group node directly on center, values under group
            group_id    = f'group__{field}'
            group_label = rel.get('group_label', rel['label'] + 's')
            add_node({'id': group_id, 'label': group_label, 'field_label': rel['label'],
                      'type': 'group', 'color': rel.get('color', 'property'),
                      'field': field})
            add_link(hostname, group_id)
            for v in values:
                node_id     = f'prop__{field}__{v}'
                child_count = Server.objects.filter(**{field: v}).values('SERVER_ID').distinct().count()
                add_node({'id': node_id, 'label': v, 'field_label': rel['label'],
                          'type': 'property', 'color': rel.get('color', 'property'),
                          'field': field, 'value': v, 'child_count': child_count})
                add_link(group_id, node_id)

    # Hosted VMs aggregate
    hv_cfg = config.get('hosted_vms')
    if hv_cfg:
        query_field = hv_cfg['query_field']
        max_list    = hv_cfg.get('max_list', 10)
        vms_qs      = Server.objects.filter(**{query_field: hostname}).values('SERVER_ID').distinct()
        vm_count    = vms_qs.count()
        if vm_count > 0:
            add_node({
                'id':          f'vms__{hostname}',
                'label':       f'{vm_count} VM{"s" if vm_count > 1 else ""}',
                'field_label': hv_cfg.get('label', 'VMs hosted'),
                'type':        'vms', 'color': 'vms',
                'total':       vm_count,
                'vm_list':     [v['SERVER_ID'] for v in vms_qs[:max_list]],
                'has_more':    vm_count > max_list,
                'field':       query_field,
                'value':       hostname,                
            })
            add_link(hostname, f'vms__{hostname}')

    return JsonResponse({'nodes': nodes, 'links': links, 'center': hostname})


@login_required
@require_http_methods(["GET"])
def property_relations(request):
    """Return servers that share a given field=value (property-centric view)."""
    MAX_SERVERS = 15
    config = _load_relation_config()

    field = request.GET.get('field', '').strip()
    value = request.GET.get('value', '').strip()
    if not field or not value:
        return JsonResponse({'error': 'Missing field or value'}, status=400)

    # Resolve human label from config
    field_label = field
    for rel in config.get('relations', []):
        if rel['field'] == field:
            field_label = rel['label']
            break

    center_id = f'prop__{field}__{value}'
    qs        = Server.objects.filter(**{field: value}).values('SERVER_ID').distinct()
    total     = qs.count()

    nodes = [{
        'id':          center_id,
        'label':       value,
        'field_label': field_label,
        'type':        'property_center',
        'color':       next((r.get('color', 'property') for r in config.get('relations', []) if r['field'] == field), 'property'),
        'field':       field,
        'value':       value,
        'total':       total,
    }]
    links = []

    for row in qs[:MAX_SERVERS]:
        sid   = row['SERVER_ID']
        count = Server.objects.filter(SERVER_ID=sid).count()
        nodes.append({
            'id':       sid,
            'label':    sid,
            'type':     'server_satellite',
            'color':    'center',
            'count':    count,
            'pivot_id': sid,
        })
        links.append({'source': center_id, 'target': sid})

    return JsonResponse({
        'nodes':    nodes,
        'links':    links,
        'center':   center_id,
        'total':    total,
        'has_more': total > MAX_SERVERS,
    })
    
    
@login_required
@require_http_methods(["GET"])
def api_column_data(request):
    """
    API endpoint for lazy loading column data

    Parameters:
        - columns: comma-separated list of column names to fetch
        - hostnames: comma-separated list of SERVER_IDs (optional, for current page)
        - page: page number (if hostnames not provided)
        - page_size: items per page
        - filters: JSON-encoded filters
        - permanentfilter: permanent filter name
        - view: 'flat' or 'grouped'

    Returns:
        JSON with column data for requested servers
    """
    columns = request.GET.get('columns', '').split(',')
    columns = [c.strip() for c in columns if c.strip()]

    if not columns:
        return JsonResponse({'error': 'No columns specified'}, status=400)

    # Validate columns against allowed fields
    json_data = get_field_labels()
    valid_fields = set(json_data.get('fields', {}).keys())
    invalid_columns = [c for c in columns if c not in valid_fields]
    if invalid_columns:
        return JsonResponse({'error': f'Invalid columns: {invalid_columns}'}, status=400)

    # Get hostnames (either directly or via pagination)
    hostnames = request.GET.get('hostnames', '').split(',')
    hostnames = [h.strip() for h in hostnames if h.strip()]

    if not hostnames:
        # Need to compute from filters and pagination
        filters_json = request.GET.get('filters', '{}')
        try:
            filters = json.loads(filters_json)
        except json.JSONDecodeError:
            filters = {}

        permanent_filter_selection = request.GET.get('permanentfilter', 'All Servers')
        page = int(request.GET.get('page', 1))
        page_size = int(request.GET.get('page_size', 50))
        flat_view = request.GET.get('view', '').lower() == 'flat'

        # Build queryset
        permanent_filter_query, _, _ = create_permanent_filter_query(json_data, permanent_filter_selection)
        queryset = Server.objects.all()

        if permanent_filter_query:
            queryset = queryset.filter(permanent_filter_query)

        # Apply filters
        su_fields = get_serverunique_fields(json_data)
        for key, values in filters.items():
            if values and key != 'ANNOTATION':
                if isinstance(values, str):
                    values = values.split(',')
                query = construct_query(key, values)
                if key in su_fields:
                    matching = ServerUnique.objects.filter(query).values_list('hostname', flat=True)
                    queryset = queryset.filter(SERVER_ID__in=matching)
                else:
                    queryset = queryset.filter(query)

        queryset = queryset.order_by('SERVER_ID', 'APP_NAME_VALUE')
        hostnames_qs = queryset.values('SERVER_ID').distinct().order_by('SERVER_ID')
        paginator = Paginator(hostnames_qs, page_size)
        page_obj = paginator.get_page(page)
        hostnames = [item['SERVER_ID'] for item in page_obj]

    if not hostnames:
        return JsonResponse({'data': {}})

    # Determine which columns are BC fields vs inventory fields
    su_fields = get_serverunique_fields(json_data)
    su_columns = [c for c in columns if c in su_fields]
    inventory_columns = [c for c in columns if c not in su_fields]

    # Fetch only the requested inventory columns
    fields_to_fetch = ['SERVER_ID'] + inventory_columns

    # Use values() for efficient column selection
    servers_data = Server.objects.filter(
        SERVER_ID__in=hostnames
    ).values(*fields_to_fetch)

    # Group by hostname
    result = defaultdict(lambda: {'instances': []})

    for server in servers_data:
        hostname = server['SERVER_ID']
        instance_data = {col: server.get(col, '') or '' for col in inventory_columns}
        result[hostname]['instances'].append(instance_data)

    # Enrich with ServerUnique data for BC columns
    if su_columns:
        su_dict = {su.hostname: su for su in ServerUnique.objects.filter(hostname__in=hostnames)}
        for hostname, data in result.items():
            su = su_dict.get(hostname)
            su_values = {col: getattr(su, col, '') or '' for col in su_columns} if su else {col: '' for col in su_columns}
            for inst in data['instances']:
                inst.update(su_values)
            data.setdefault('constant_fields', {}).update(su_values)

    # Add summary info
    summaries = ServerGroupSummary.objects.filter(
        SERVER_ID__in=hostnames
    ).only('SERVER_ID', 'constant_fields', 'variable_fields')

    for summary in summaries:
        hostname = summary.SERVER_ID
        if hostname in result:
            result[hostname]['constant_fields'] = {
                col: summary.constant_fields.get(col, '')
                for col in columns if col in summary.constant_fields
            }
            result[hostname]['variable_fields'] = {
                col: summary.variable_fields.get(col, {})
                for col in columns if col in summary.variable_fields
            }

    return JsonResponse({
        'data': dict(result),
        'columns': columns,
        'hostnames': hostnames
    })


@login_required
@require_http_methods(["GET"])
def api_listbox_values(request):
    # API endpoint to get listbox values for specific column for lazy loading dropdown options
    columns = request.GET.get('columns', '').split(',')
    columns = [c.strip() for c in columns if c.strip()]

    if not columns:
        return JsonResponse({'error': 'No columns specified'}, status=400)

    # Validate columns
    json_data = get_field_labels()
    valid_fields = set(json_data.get('fields', {}).keys())
    columns = [c for c in columns if c in valid_fields]

    # Batch load listbox values
    su_fields = get_serverunique_fields(json_data)
    listbox_values = batch_load_listbox_values(columns, su_fields=su_fields)

    return JsonResponse({'data': listbox_values})


def get_filter_mapping():
    """
    Optimized version - uses cached JSON config
    """
    json_data = get_field_labels()
    return {field_name: field_name for field_name in json_data['fields'].keys()}


@login_required
@require_http_methods(["GET"])
def api_server_count(request):
    """
    API endpoint to get server count without loading data
    Used for quick statistics display

    Returns:
        {
            "total_instances": 15000,
            "unique_servers": 12500
        }
    """
    json_data = get_field_labels()

    # Get permanent filter
    permanent_filter_selection = request.GET.get('permanentfilter', 'All Servers')

    # Base queryset
    queryset = Server.objects.all()

    # Apply permanent filter
    permanent_filter_query, _, _ = create_permanent_filter_query(json_data, permanent_filter_selection)
    if permanent_filter_query:
        queryset = queryset.filter(permanent_filter_query)

    # Apply user filters
    for field_key, field_info in json_data.get('fields', {}).items():
        input_name = field_info.get('inputname')
        if input_name:
            filter_value = request.GET.get(input_name, '')
            if filter_value:
                values = [v.strip() for v in filter_value.split(',') if v.strip()]
                if values:
                    query = construct_query(field_key, values)
                    queryset = queryset.filter(query)

    # Count (optimized)
    total_instances = queryset.count()
    unique_servers = queryset.values('SERVER_ID').distinct().count()

    return JsonResponse({
        'total_instances': total_instances,
        'unique_servers': unique_servers
    })


@login_required
def chart_view(request):
    json_data = get_field_labels()
    
    selected_fields = request.GET.getlist('fields')
    chart_types = request.GET.getlist('types')
    permanent_filter_selection = request.GET.get('permanentfilter')
    
    requestfilters = {}
    for key, value in request.GET.items():
        requestfilters[key] = value
    
    servers = get_filtered_servers(requestfilters, permanent_filter_selection)

    su_fields = get_serverunique_fields(json_data)
    su_selected = [f for f in selected_fields if f in su_fields]
    non_su_selected = [f for f in selected_fields if f not in su_fields and f != 'ANNOTATION']
    
    if "ANNOTATION" in selected_fields:
        server_ids = list(servers.values_list('SERVER_ID', flat=True).distinct())
        annotations = ServerAnnotation.objects.filter(SERVER_ID__in=server_ids)
        annotation_map = {
            ann.SERVER_ID: ' | '.join(
                f"[{e.get('type') or '?'}] {e.get('text') or ''}" for e in ann.active_annotations
            ) or (ann.notes or '')
            for ann in annotations
        }
        
        fields_to_extract = ['SERVER_ID'] + non_su_selected

        server_data = list(servers.values(*fields_to_extract).distinct())
        
        for server in server_data:
            server['ANNOTATION'] = annotation_map.get(server['SERVER_ID'], '')
    else:
        fields_to_extract = ['SERVER_ID'] + non_su_selected
        server_data = list(servers.values(*fields_to_extract).distinct())

    # Enrich with ServerUnique data for BC fields
    if su_selected:
        all_hostnames = list(set(s['SERVER_ID'] for s in server_data))
        su_dict = {su.hostname: su for su in ServerUnique.objects.filter(hostname__in=all_hostnames)}
        for server in server_data:
            su = su_dict.get(server['SERVER_ID'])
            for f in su_selected:
                server[f] = getattr(su, f, '') or '' if su else ''
    
    # Pre-calculate the fields total here
    field_totals = {}
    for field in selected_fields:
       
        # Count the unique combos (SERVER_ID, valeur)
        unique_combos = set()
        for server in server_data:
            server_id = server.get('SERVER_ID')
            value = server.get(field, 'Unknown')
            if value is None or value == '':
                value = 'Unknown'
            unique_combos.add((server_id, str(value)))
        
        field_totals[field] = len(unique_combos)
    
    return generate_charts(request, server_data, json_data, selected_fields, chart_types, field_totals)

# View to load a saved search and redirect to the search results page
@login_required        
def load_search(request, search_id):

    # Retrieve the saved search object for the given search_id and user
    saved_search = SavedSearch.objects.get(id=search_id, user_profile__user=request.user)
    
    # Get the filters from the saved search
    filters = saved_search.filters
    
    # Convert the filters dictionary to a query string
    query_string = urlencode(filters, doseq=True)

    return redirect(f"/{app_name}/?{query_string}")


# View to save a user's search criteria
@login_required
def save_search(request):

    # filters, tags, referrer, query and search_name are passed through saveSearchForm
    if request.method == 'POST':
    
        search_name = request.POST['search_name']  # Retrieve the search name from the POST data
        if len(search_name) > 25:  # Validate the name length
            return JsonResponse({'success': False, 'message': 'Search name must be 25 characters or fewer.'})

        profile = UserProfile.objects.get(user=request.user)  # Get the user's profile
        # Check if a search with the same name already exists for the user
        if SavedSearch.objects.filter(user_profile=profile, name=search_name).exists():
            return JsonResponse({'success': False, 'message': 'A search with this name already exists.'})

        # Retrieve the filters from the POST data        
        filters = request.POST.get('filters', None)
        if not filters:
            return JsonResponse({'success': False, 'message': 'Filters cannot be empty.'})

        try:
            search_params = json.loads(filters)  # Parse the filters data, passed as JSON
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid filter data.'})

        tags = json.loads(request.POST.get('tags', '[]'))  # Retrieve the tags from the POST data, if any
        # Create a new SavedSearch entry with the provided data        
        saved_search = SavedSearch.objects.create(user_profile=profile, name=search_name, filters=search_params, tags=tags, view=app_name)

        # Redirect back to the previous page after saving the search
        return redirect(request.META.get('HTTP_REFERER', '/'))
        
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


# View to delete a saved search
@login_required
def delete_search(request, search_id):

    # Retrieve the saved search object for the given search_id and user
    saved_search = SavedSearch.objects.get(id=search_id, user_profile__user=request.user)

    # Delete the saved search object
    saved_search.delete()

    # Redirect back to the previous page after deleting the search
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def log_snapshot_status(request):
    if not has_perm(request.user, 'inventory.view'):
        return redirect('login')
    logs = SnapshotStatus.objects.order_by('-date_import')[:100]
    return render(request, 'inventory/logs_snapshot.html', {'logs': logs})


# Display the logs_imports.html with the last 100 logs entries
def log_imports(request):

    # Get and sort the last entries of the table inventory_importstatus
    logs = ImportStatus.objects.order_by('-date_import')[:100]

    return render(request, f"{app_name}/logs_imports.html", {
        "logs": logs
    })
    

# Filters servers based on the provided criteria and applies a permanent filter if selected
def get_filtered_servers(requestfilters, permanent_filter_selection):

    json_data = get_field_labels()
    su_fields = get_serverunique_fields(json_data)

    filters = {}

    for field_name, field_properties in json_data['fields'].items():
        input_name = field_properties.get('inputname')
        if not input_name:
            continue
        if field_properties.get('fieldtype') == 'date':
            date_from = (requestfilters.get(f'{input_name}_from') or '').strip()
            date_to = (requestfilters.get(f'{input_name}_to') or '').strip()
            if date_from:
                filters[f'{field_name}__gte'] = [date_from]
            if date_to:
                filters[f'{field_name}__lte'] = [date_to]
        else:
            value = requestfilters.get(input_name, None)
            if value not in ('', None):
                filters[field_name] = value

    servers = Server.objects.all().order_by('SERVER_ID')
    su_filter_query = Q()
    combined_filter_query = Q()

    for key, value in filters.items():
        if key.endswith('__gte') or key.endswith('__lte'):
            field_name, lookup = key.rsplit('__', 1)
            combined_filter_query &= Q(**{f'{field_name}__{lookup}': value[0]})
        else:
            terms = value.split(',') if isinstance(value, str) and ',' in value else [value]
            query = construct_query(key, terms)
            if key in su_fields:
                su_filter_query &= query
            else:
                combined_filter_query &= query

    if combined_filter_query:
        servers = servers.filter(combined_filter_query)

    if su_filter_query:
        matching_hostnames = ServerUnique.objects.filter(su_filter_query).values_list('hostname', flat=True)
        servers = servers.filter(SERVER_ID__in=matching_hostnames)

    permanent_filter_query, permanent_filter_names, permanent_filter_attributes = create_permanent_filter_query(json_data, permanent_filter_selection)
    if permanent_filter_query:
        servers = servers.filter(permanent_filter_query)

    return servers


# Handles the export of server data to a specified file format (CSV or XLSX)
def export_to_file(request, filetype):

    # filters, columns and permanentfilterselection are passed as data content
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not authorized'}, status=405)
        
    # Verify that the file type is allowed
    filetype = filetype.lower()
    if filetype not in ['xlsx', 'csv']:
        return JsonResponse({'error': 'Invalid format type'}, status=400)

    # Get the data passed as content
    data = json.loads(request.body)
    requestfilters = data['filters']
    columns = data['columns']
    permanent_filter_selection = data['permanentfilterselection']
    exportnotes = 'ANNOTATION' in columns
    
    # Filter the servers based on the filters and permanent filter
    servers = get_filtered_servers(requestfilters, permanent_filter_selection)
    
    hostnames = (servers.values('SERVER_ID').distinct().order_by('SERVER_ID'))
    hostname_list = [item['SERVER_ID'] for item in hostnames]
    
    annotations_dict = {}
    if exportnotes:
        if hostname_list:
            annotations = ServerAnnotation.objects.filter(SERVER_ID__in=hostname_list)
            annotations_dict = {
                ann.SERVER_ID: ' | '.join(
                    f"[{e.get('type') or '?'}] {e.get('text') or ''}" for e in ann.active_annotations
                ) or (ann.notes or '')
                for ann in annotations
            }

    # Build ServerUnique lookup for BC fields in export
    json_data = get_field_labels()
    su_fields = get_serverunique_fields(json_data)
    su_columns = [c for c in columns if c in su_fields]
    su_dict = {}
    if su_columns:
        su_dict = {su.hostname: su for su in ServerUnique.objects.filter(hostname__in=hostname_list)}

    # Generate the filter mapping
    FILTER_MAPPING=get_filter_mapping()
    job_id = str(uuid.uuid4())  # Generate a unique identifier
    filepath = os.path.join(EXPORT_DIR, f"{job_id}.{filetype}")
    
    # Generate the export file using generate_excel from the exports.py module
    def background_export():
        try:
            if filetype == 'xlsx':
                generate_excel(filepath, servers, annotations_dict, columns, FILTER_MAPPING, su_dict=su_dict, su_fields=su_fields)
            else:
                generate_csv(filepath, servers, annotations_dict, columns, FILTER_MAPPING, su_dict=su_dict, su_fields=su_fields)
        except Exception as e:
            print(f"Error during the export : {e}")

    # Execute the fucntion background_export as a separated thread
    threading.Thread(target=background_export).start()

    # Return the job id back in order for the js function startExport to wait for its completion
    return JsonResponse({'job_id': job_id})


# Handles the export of grouped server data to a specified file format (CSV or XLSX)
def export_to_file_grouped(request, filetype):

    # filters, columns and permanentfilterselection are passed as data content
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not authorized'}, status=405)
        
    # Verify that the file type is allowed
    filetype = filetype.lower()
    if filetype not in ['xlsx', 'csv']:
        return JsonResponse({'error': 'Invalid format type'}, status=400)

    # Get the data passed as content
    data = json.loads(request.body)
    requestfilters = data['filters']
    columns = data['columns']
    permanent_filter_selection = data['permanentfilterselection']
    exportnotes = 'ANNOTATION' in columns
    
    # Filter the servers based on the filters and permanent filter
    servers = get_filtered_servers(requestfilters, permanent_filter_selection)
    
    hostnames = (servers.values('SERVER_ID').distinct().order_by('SERVER_ID'))
    server_groups = defaultdict(list)
    for server in servers.order_by('SERVER_ID'):
        server_groups[server.SERVER_ID].append(server)
    
    hostname_list = [item['SERVER_ID'] for item in hostnames]
    summaries_dict = {}
    if hostname_list:
        summaries = ServerGroupSummary.objects.filter(SERVER_ID__in=hostname_list)
        summaries_dict = {s.SERVER_ID: s for s in summaries}
        
    annotations_dict = {}
    if exportnotes:
        if hostname_list:
            annotations = ServerAnnotation.objects.filter(SERVER_ID__in=hostname_list)
            annotations_dict = {
                ann.SERVER_ID: ' | '.join(
                    f"[{e.get('type') or '?'}] {e.get('text') or ''}" for e in ann.active_annotations
                ) or (ann.notes or '')
                for ann in annotations
            }

    # Build ServerUnique lookup for BC fields in export
    json_data = get_field_labels()
    su_fields = get_serverunique_fields(json_data)
    su_columns = [c for c in columns if c in su_fields]
    su_dict = {}
    if su_columns:
        su_dict = {su.hostname: su for su in ServerUnique.objects.filter(hostname__in=hostname_list)}
    
    # Generate the filter mapping
    job_id = str(uuid.uuid4())  # Generate a unique identifier
    filepath = os.path.join(EXPORT_DIR, f"{job_id}.{filetype}")
    
    # Generate the export file using generate_excel from the exports.py module
    def background_export():
        try:
            if filetype == 'xlsx':
                generate_excel_grouped(filepath, hostnames, server_groups, summaries_dict, annotations_dict, columns, su_dict=su_dict, su_fields=su_fields)
            else:
                generate_csv_grouped(filepath, hostnames, server_groups, summaries_dict, annotations_dict, columns, su_dict=su_dict, su_fields=su_fields)
        except Exception as e:
            print(f"Error during the export : {e}")

    # Execute the fucntion background_export as a separated thread
    threading.Thread(target=background_export).start()

    # Return the job id back in order for the js function startExport to wait for its completion
    return JsonResponse({'job_id': job_id})


# Check the state of the file to export 
def export_status(request, job_id, filetype):

    extension = filetype.lower()
    if extension not in ['xlsx', 'csv']:
        return JsonResponse({'Error': 'Invalid filetype'}, status=400)

    # Create the full file name based on the job_id and extension passed in the URL from the s function startExport
    filename=f'{job_id}.{extension}'
    filepath = os.path.join(EXPORT_DIR, filename)
    
    # Return "ready" id the file is present, and "pending" if not. During its creation, the file as an additional .tmp extension to not trigger a "ready" too early
    if os.path.exists(filepath):
        return JsonResponse({'status': 'ready'})
    return JsonResponse({'status': 'pending'})


# Trigger the automatic downlaod when the export file is fully created
def download_export(request, job_id, filetype):

    extension = filetype.lower()
    if extension not in ['xlsx', 'csv']:
        return JsonResponse({'Error': 'Invalid filetype'}, status=400)
        
    # Create the full file name based on the job_id and extension passed in the URL from the s function startExport
    filename=f'{job_id}.{extension}'
    filepath = os.path.join(EXPORT_DIR, filename)
    
    # Raise an error if the file doesn't exist
    if not os.path.exists(filepath):
        raise Http404("File not found")

    # Initiate the automatic download
    response = FileResponse(open(filepath, 'rb'), as_attachment=True, filename=f'export_{app_name}.{extension}')

    def delete_file_after_close(resp):
        try:
            os.remove(filepath)
        except Exception as e:
            print(f"Error during the removing of the file : {e}")

    # Close and delete the generated file once downloaded
    response.close = lambda *args, **kwargs: delete_file_after_close(response)
    return response
        


# Update and save the permanent filter field for the authenticated user
def update_permanentfilter_field(request):
    if request.method == "POST":
        
        # Get the permanent filter choice from the POST data, passed as input permanentfilter_choice
        permanentfilter_choice = request.POST.get("permanentfilter_choice")

        if request.user.is_authenticated:
            # Get the user's profile
            profile = UserProfile.objects.get(user=request.user)

            # Update or create the SavedOptions entry for the user
            permanentfilter, created = SavedOptions.objects.update_or_create(  # Update or create the SavedOptions entry for the user
                user_profile=profile,
                defaults={f'{app_name}_permanentfilter': permanentfilter_choice}
            )
            # Redirect back to the referring page after submission
            return redirect(request.META.get('HTTP_REFERER', '/'))  # Redirect back to the same page after submission
        else:
            # The user is not authenticated
            return JsonResponse({"status": "error", "message": "User is not authenticated."}, status=401)

    # Return an error message if the request method is invalid
    return JsonResponse({"status": "error", "message": "Invalid request method."}, status=400)
    
    
@login_required
@require_http_methods(["GET", "POST"])
def edit_annotation(request, hostname):
    annotation = ServerAnnotation.objects.filter(SERVER_ID=hostname).first()

    if request.method == 'GET':
        form = AnnotationForm(instance=annotation)

        return JsonResponse({
            'hostname': hostname,
            'notes': annotation.notes if annotation else '',
            'type': annotation.type if annotation else '',
            'servicenow': annotation.servicenow if annotation else '',
            'history': annotation.get_history_display() if annotation else [],
            'active_annotations': annotation.active_annotations if annotation else [],
            'form_html': form.as_p()
        })

    elif request.method == 'POST':
        action = request.POST.get('action', 'add')

        if action == 'resolve':
            try:
                entry_index = int(request.POST.get('entry_index', -1))
            except (ValueError, TypeError):
                return JsonResponse({'success': False, 'message': 'Invalid entry index'}, status=400)

            if not annotation or not annotation.history or entry_index < 0 or entry_index >= len(annotation.history):
                return JsonResponse({'success': False, 'message': 'Annotation entry not found'}, status=404)

            annotation.resolve_entry(entry_index, request.user)
            return JsonResponse({
                'success': True,
                'message': 'Annotation resolved',
                'active_annotations': annotation.active_annotations,
                'history': annotation.get_history_display(),
                'notes': annotation.notes,
                'type': annotation.type,
                'servicenow': annotation.servicenow,
            })

        elif action == 'clear':
            if annotation and annotation.history:
                now_iso = timezone.now().isoformat()
                username = request.user.username
                changed = False
                for i, entry in enumerate(annotation.history):
                    if entry.get('is_active', True):
                        annotation.history[i]['is_active'] = False
                        annotation.history[i]['resolved_at'] = now_iso
                        annotation.history[i]['resolved_by'] = username
                        changed = True
                if changed:
                    annotation._sync_from_active()
                    annotation.save()
            return JsonResponse({
                'success': True,
                'message': 'All annotations cleared',
                'active_annotations': [],
                'history': annotation.get_history_display() if annotation else [],
                'notes': '',
                'type': '',
                'servicenow': '',
            })

        # action == 'add' (default)
        form = AnnotationForm(request.POST, instance=annotation)

        if form.is_valid():
            notes_text = form.cleaned_data['notes'].strip()
            annotation_type = form.cleaned_data['type']
            custom_type = form.cleaned_data.get('custom_type', '').strip()
            servicenow = form.cleaned_data['servicenow'].strip()

            # Use custom type if "CUSTOM" is selected
            if annotation_type == 'CUSTOM':
                annotation_type = custom_type

            if not annotation:
                annotation = ServerAnnotation(SERVER_ID=hostname)

            annotation.notes = notes_text
            annotation.type = annotation_type
            annotation.servicenow = servicenow

            annotation.add_entry(notes_text, request.user, annotation_type, servicenow)
            annotation.save()

            return JsonResponse({
                'success': True,
                'message': 'Annotation saved successfully',
                'notes': annotation.notes,
                'type': annotation.type,
                'servicenow': annotation.servicenow,
                'active_annotations': annotation.active_annotations,
                'history': annotation.get_history_display(),
                'updated_by': request.user.username,
                'updated_at': annotation.updated_at.strftime('%d/%m/%Y %H:%M')
            })
        
        
# Handles the bulk import of server data from a CSV file
@require_http_methods(["GET", "POST"])
def bulk_import_csv(request):
    if request.method == 'POST':
        form = CsvUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            if not csv_file.name.endswith('.csv'):
                return JsonResponse({'error': "This file must be a CSV."}, status=400)

            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, delimiter=';')

            base_headers = ['hostname', 'type', 'servicenow', 'notes']
            extended_headers = base_headers + ['action']

            errors = []
            if reader.fieldnames not in (base_headers, extended_headers):
                errors.append(
                    f"The Header is incorrect. Fields expected: {', '.join(base_headers)} "
                    f"(optionally with 'action' column). "
                    f"Received: {', '.join(reader.fieldnames or [])}"
                )
                return JsonResponse({'error': 'Header incorrect.', 'detailed_errors': errors}, status=400)

            has_action_col = (reader.fieldnames == extended_headers)
            header_len = len(reader.fieldnames)
            for idx, line in enumerate(decoded_file[1:], start=2):
                if line.count(';') + 1 != header_len:
                    errors.append(
                        f"Line {idx}: number of columns is incorrect "
                        f"(Found: {line.count(';') + 1}, expected: {header_len})."
                    )

            if errors:
                return JsonResponse({'error': f"{len(errors)} errors detected.", 'detailed_errors': errors[:50]}, status=400)

            updated_count = 0
            ignored_count = 0
            lines_processed = 0
            total_lines = len(decoded_file) - 1
            errors = []
            instances_to_update = []
            now_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            now_usr = request.user.username
            now_str = f"[{now_date} - {now_usr}]"

            def process_import():
                nonlocal updated_count, ignored_count, lines_processed, total_lines, errors
                yield f"data: {json.dumps({'progress': 0, 'message': 'Reading CSV file...'})}\n\n"

                for row in reader:
                    hostname = row.get('hostname')

                    if not hostname:
                        ignored_count += 1
                        errors.append(f"Line ignored: 'hostname' is missing. Line content: {row}")
                        continue

                    if not Server.objects.filter(SERVER_ID=hostname).exists():
                        ignored_count += 1
                        errors.append(f"Line ignored: 'hostname'={hostname} does not exist as a server.")
                        continue

                    try:
                        annotation, created = ServerAnnotation.objects.get_or_create(SERVER_ID=hostname)
                    except ServerAnnotation.DoesNotExist:
                        # This should never happen, since we're using get_or_create
                        errors.append(f"Line ignored: unable to retrieve or create 'hostname'={hostname}.")
                        continue

                    csv_action = row.get('action', 'add').strip().lower() if has_action_col else 'add'

                    if csv_action == 'resolve':
                        active = [(i, e) for i, e in enumerate(annotation.history or []) if e.get('is_active', True)]
                        if active:
                            now_iso = timezone.now().isoformat()
                            latest_idx = max(active, key=lambda x: x[1].get('date', ''))[0]
                            annotation.history[latest_idx]['is_active'] = False
                            annotation.history[latest_idx]['resolved_at'] = now_iso
                            annotation.history[latest_idx]['resolved_by'] = now_usr
                            annotation._sync_from_active()
                            instances_to_update.append(annotation)
                        else:
                            ignored_count += 1
                            errors.append(f"Line ignored: no active annotations for '{hostname}'.")
                            continue
                    else:
                        notes = row.get('notes', "") or None
                        xtype = row.get('type', "").upper() or None
                        servicenow = row.get('servicenow', "") or None

                        if notes is not None:
                            annotation.notes = notes
                        if xtype is not None:
                            annotation.type = xtype
                        if servicenow is not None:
                            annotation.servicenow = servicenow

                        annotation.add_entry(notes, request.user, xtype, servicenow)
                        instances_to_update.append(annotation)

                    lines_processed += 1
                    progress = (lines_processed / total_lines) * 100
                    yield f"data: {json.dumps({'progress': progress, 'message': 'Reading CSV file...'})}\n\n"

                total_updates = len(instances_to_update)
                yield f"data: {json.dumps({'progress': 0, 'message': 'Importing data...'})}\n\n"
                for i in range(0, total_updates, 100):
                    ServerAnnotation.objects.bulk_update(
                        instances_to_update[i:i+100],
                        ['notes', 'type', 'servicenow', 'history']
                    )
                    updated_count += len(instances_to_update[i:i+100])
                    progress = (updated_count / total_updates) * 100
                    yield f"data: {json.dumps({'progress': progress, 'message': 'Importing data...'})}\n\n"

                success_msg = f"Import successful: {updated_count} servers updated. {ignored_count} servers ignored."
                yield f"data: {json.dumps({'success': success_msg})}\n\n"

                if errors:
                    warning_msg = f"{len(errors)} warnings detected."
                    yield f"data: {json.dumps({'warning': warning_msg, 'detailed_errors': errors[:50]})}\n\n"

            response = StreamingHttpResponse(process_import(), content_type='text/event-stream')
            response['Cache-Control'] = 'no-cache'
            return response

        error_msg = "Form validation failed or file is incorrect."
        return JsonResponse({'error': error_msg}, status=400)

    elif request.method == 'GET':
        form = CsvUploadForm()
        return render(request, f"{app_name}/bulk_import.html", {'form': form})

    return JsonResponse({'error': 'Invalid request method'}, status=405)
    

# View to handle bulk update of servers and providing progress updates
def servers_bulk_update(request):
    if request.method == "POST":
        try:
            # Retrieve the filters from the POST data
            query_param = request.POST.get('query')

            # Check if the query_param exists
            if not query_param:
                return JsonResponse({'status': 'error', 'message': 'Query string not provided'}, status=400)

            # Decode and parse the query string
            query_param = unquote(query_param)
            parsed_query = parse_qs(query_param)

            # Get the user's profile
            profile = UserProfile.objects.get(user=request.user)

            # Read and save the field_labels.json information
            json_path = os.path.join(os.path.dirname(__file__), 'field_labels.json')
            with open(json_path, "r", encoding="utf-8") as f:
                json_data = json.load(f)

            # Get the permanent filter choice from the POST data
            permanent_filter_selection = request.POST.get("permanentfilter_choice")

            # Define servers_to_bulkedit to first get all objects
            servers_to_bulkedit = Server.objects.all().order_by('SERVER_ID')

            # Call create_permanent_filter_query to create the query associated to the permanent filter parameters
            permanent_filter_query, permanent_filter_names, permanent_filter_attributes = create_permanent_filter_query(json_data, permanent_filter_selection)
            if permanent_filter_query:
                servers_to_bulkedit = servers_to_bulkedit.filter(permanent_filter_query)

            # Initialize the filters dictionary
            filters = {}

            # Create the filter using the parameter passed in the URL with request.GET.get
            # Iterate over the fields section to search for potential matches and construct the filters dictionary
            for field_key, field_info in json_data['fields'].items():
                if 'inputname' in field_info:
                    filters[field_key] = parsed_query.get(field_info['inputname'], [''])[0].split(',')

            # Remove all [''] added during the filters creation
            filters = {k: v for k, v in filters.items() if v != ['']}

            # Build the query based on the filters values extracted earlier
            for key, value in filters.items():
                if isinstance(value, list):
                    terms = value
                else:
                    terms = [value]
                query = construct_query(key, terms)
                servers_to_bulkedit = servers_to_bulkedit.filter(query)

            # Retrieve the servers to update: get the unique server hostnames on the objects filtered with filters
            total_updates = servers_to_bulkedit.count()

            if total_updates == 0:
                return JsonResponse({'status': 'warning', 'message': 'No servers to update with the given filters'}, status=200)

            # Collect input data for updates and convert to uppercase if needed
            bulk_type = request.POST.get('bulk_type', '')
            bulk_custom_type = request.POST.get('bulk_custom_type', '')
            bulk_servicenow = request.POST.get('bulk_servicenow', '')
            bulk_notes = request.POST.get('bulk_notes', '')
            if bulk_type == "CUSTOM":
                bulk_type = bulk_custom_type

            # Function to handle the bulk update of server records in a way that provides progress updates
            def update_with_progress():
                batch_size = 50
                total_batches = (total_updates + batch_size - 1) // batch_size

                # Iterate over each batch
                for i in range(total_batches):
                    batch_start = i * batch_size
                    batch_end = min((i + 1) * batch_size, total_updates)
                    batch = servers_to_bulkedit[batch_start:batch_end]

                    # Get the corresponding ServerAnnotation instances
                    annotations = []
                    for server in batch:
                        annotation, created = ServerAnnotation.objects.get_or_create(SERVER_ID=server.SERVER_ID)
                        annotations.append(annotation)

                    # Update the annotations
                    for annotation in annotations:
                        # Add entry to history
                        annotation.add_entry(bulk_notes, request.user, bulk_type, bulk_servicenow)
                        annotation.notes = bulk_notes
                        annotation.type = bulk_type
                        annotation.servicenow = bulk_servicenow

                    # Perform a bulk update on the current batch of records
                    ServerAnnotation.objects.bulk_update(annotations, ['notes', 'type', 'servicenow', 'history'])

                    progress = ((i + 1) / total_batches) * 100
                    yield f"data: progress:{progress}|batch:{i + 1}\n\n"

            response = StreamingHttpResponse(update_with_progress(), content_type='text/event-stream')
            response['Cache-Control'] = 'no-cache'
            return response

        except Exception as e:
            # Log unexpected errors and return a meaningful message
            return JsonResponse({'status': 'error', 'message': f"An unexpected error occurred: {e}"}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Invalid Request Method'}, status=400)


@csrf_exempt
@require_http_methods(["PUT"])
def api_bulk_annotation(request):
    """
    API endpoint to bulk upsert annotations.

    PUT /inventory/api/annotations/
    Authorization: Token <token>
    Content-Type: application/json

    Body: JSON array:
      [{"SERVER_ID": "...", "notes": "...", "type": "...", "servicenow": "..."}, ...]

    SERVER_ID is required. Other fields default to ''.

    Response: { "success": true, "created": N, "updated": N, "errors": [...] }
    """
    from rest_framework.authtoken.models import Token

    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if not auth_header.startswith('Token '):
        return JsonResponse({'success': False, 'error': 'Authentication required'}, status=401)

    token_key = auth_header[6:]
    try:
        token = Token.objects.select_related('user').get(key=token_key)
    except Token.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid token'}, status=401)

    user = token.user
    if not user.is_active:
        return JsonResponse({'success': False, 'error': 'User inactive'}, status=401)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON body'}, status=400)

    if not isinstance(data, list):
        return JsonResponse({'success': False, 'error': 'Expected a JSON array'}, status=400)

    if not data:
        return JsonResponse({'success': False, 'error': 'Empty list'}, status=400)

    errors = []
    instances_to_update = []
    created_count = 0
    updated_count = 0
    now = timezone.now().isoformat()

    # Resolve canonical SERVER_IDs (case-insensitive) in one query
    from django.db.models.functions import Lower
    input_lowers = {
        (item.get('SERVER_ID') or '').strip().lower()
        for item in data
        if (item.get('SERVER_ID') or '').strip()
    }
    canonical_map = {
        sid.lower(): sid
        for sid in Server.objects
            .annotate(_lower=Lower('SERVER_ID'))
            .filter(_lower__in=input_lowers)
            .values_list('SERVER_ID', flat=True)
            .distinct()
    }

    for item in data:
        server_id = (item.get('SERVER_ID') or '').strip()
        if not server_id:
            errors.append({'SERVER_ID': '', 'error': 'SERVER_ID is required'})
            continue

        canonical_id = canonical_map.get(server_id.lower())
        if canonical_id is None:
            errors.append({'SERVER_ID': server_id, 'error': 'Server not found in inventory'})
            continue

        notes = (item.get('notes') or '').strip()
        annotation_type = (item.get('type') or '').strip()
        servicenow = (item.get('servicenow') or '').strip()

        annotation, created = ServerAnnotation.objects.get_or_create(SERVER_ID=canonical_id)

        if not annotation.history:
            annotation.history = []
        annotation.history.append({
            'text': notes,
            'user': user.username,
            'date': now,
            'type': annotation_type,
            'servicenow': servicenow,
        })
        annotation.notes = notes
        annotation.type = annotation_type
        annotation.servicenow = servicenow

        instances_to_update.append(annotation)
        if created:
            created_count += 1
        else:
            updated_count += 1

    batch_size = 100
    for i in range(0, len(instances_to_update), batch_size):
        ServerAnnotation.objects.bulk_update(
            instances_to_update[i:i + batch_size],
            ['notes', 'type', 'servicenow', 'history']
        )

    return JsonResponse({
        'success': True,
        'created': created_count,
        'updated': updated_count,
        'errors': errors,
    })


_FIELD_TRACKING_CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'field_tracking.json')


def _snapshot_qs_for_request(request, field_name, filter_field='', filter_value='', filter_field2='', filter_value2='', filter_field3='', filter_value3=''):
    """Return a filtered FieldSnapshotFiltered queryset based on request date params."""
    qs = FieldSnapshotFiltered.objects.filter(
        field_name=field_name,
        filter_field=filter_field,   filter_value=filter_value,
        filter_field2=filter_field2, filter_value2=filter_value2,
        filter_field3=filter_field3, filter_value3=filter_value3,
    )
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    if date_from or date_to:
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
    else:
        try:
            days = int(request.GET.get('days', 90))
        except ValueError:
            days = 90
        if days > 0:
            qs = qs.filter(date__gte=datetime.date.today() - datetime.timedelta(days=days))
    return qs


@login_required
def field_snapshots_dashboard(request):
    if not has_perm(request.user, 'inventory.view'):
        return redirect('login')
    try:
        with open(_FIELD_TRACKING_CONFIG_PATH, 'r', encoding='utf-8') as f:
            fields_config = json.load(f)
    except FileNotFoundError:
        fields_config = []

    # Build {field_name: {filters: [...], combinations: [...]}} for scope dropdowns
    filter_options: dict = {}
    for fc in fields_config:
        fn = fc['field']
        filters_list = []
        for fconf in fc.get('filters', []):
            ff = fconf['field']
            values = list(
                FieldSnapshotFiltered.objects
                .filter(field_name=fn, filter_field=ff, filter_field2='')
                .values_list('filter_value', flat=True)
                .distinct()
                .order_by('filter_value')
            )
            if values:
                filters_list.append({
                    'field': ff,
                    'label': fconf.get('label', ff),
                    'values': values,
                })
        combinations = [combo['fields'] for combo in fc.get('filter_combinations', [])]
        if filters_list:
            filter_options[fn] = {'filters': filters_list, 'combinations': combinations}

    last_snapshot = SnapshotStatus.objects.order_by('-date_import').first()

    return render(request, 'inventory/field_snapshots.html', {
        'fields_config': fields_config,
        'fields_config_json': json.dumps(fields_config),
        'filter_options_json': json.dumps(filter_options),
        'last_snapshot': last_snapshot,
    })


@login_required
def field_snapshots_data(request):
    """JSON time-series for a tracked field: {labels, datasets}."""
    if not has_perm(request.user, 'inventory.view'):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    field_name = request.GET.get('field', '').strip()
    if not field_name:
        return JsonResponse({'error': 'field parameter required'}, status=400)

    filter_field  = request.GET.get('filter_field',  '').strip()
    filter_value  = request.GET.get('filter_value',  '').strip()
    filter_field2 = request.GET.get('filter_field2', '').strip()
    filter_value2 = request.GET.get('filter_value2', '').strip()
    filter_field3 = request.GET.get('filter_field3', '').strip()
    filter_value3 = request.GET.get('filter_value3', '').strip()

    qs = _snapshot_qs_for_request(request, field_name, filter_field, filter_value, filter_field2, filter_value2, filter_field3, filter_value3)
    rows = list(qs.order_by('date').values('date', 'counts'))
    if not rows:
        return JsonResponse({'labels': [], 'datasets': []})

    date_set = sorted({r['date'] for r in rows})
    value_totals: dict = {}
    data_by_value: dict = {}

    for r in rows:
        for fv, count in r['counts'].items():
            if fv not in data_by_value:
                data_by_value[fv] = {}
                value_totals[fv] = 0
            data_by_value[fv][r['date']] = count
            value_totals[fv] += count

    value_set = sorted(value_totals, key=lambda v: value_totals[v], reverse=True)
    labels = [d.strftime('%d.%m.%Y') for d in date_set]
    datasets = [
        {'label': v or '(empty)', 'data': [data_by_value[v].get(d) for d in date_set]}
        for v in value_set
    ]
    return JsonResponse({'labels': labels, 'datasets': datasets})


@login_required
def field_snapshots_export(request):
    """Export snapshot data as Excel. all=1 exports every tracked field (one sheet each)."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
    from openpyxl.utils import get_column_letter

    if not has_perm(request.user, 'inventory.view'):
        return JsonResponse({'error': 'Forbidden'}, status=403)

    try:
        with open(_FIELD_TRACKING_CONFIG_PATH, 'r', encoding='utf-8') as f:
            fields_config = json.load(f)
    except FileNotFoundError:
        fields_config = []

    label_map = {fc['field']: fc.get('label', fc['field']) for fc in fields_config}

    if request.GET.get('all') == '1':
        field_names = [fc['field'] for fc in fields_config]
        filename = f"field_trends_all_{datetime.date.today():%Y%m%d}.xlsx"
    else:
        field_name = request.GET.get('field', '').strip()
        if not field_name:
            return JsonResponse({'error': 'field parameter required'}, status=400)
        field_names = [field_name]
        filename = f"field_trends_{field_name}_{datetime.date.today():%Y%m%d}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='4F46E5')
    center_align = Alignment(horizontal='center')
    right_align  = Alignment(horizontal='right')
    thin_border  = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
    )

    export_filter_field  = request.GET.get('filter_field',  '').strip()
    export_filter_value  = request.GET.get('filter_value',  '').strip()
    export_filter_field2 = request.GET.get('filter_field2', '').strip()
    export_filter_value2 = request.GET.get('filter_value2', '').strip()
    export_filter_field3 = request.GET.get('filter_field3', '').strip()
    export_filter_value3 = request.GET.get('filter_value3', '').strip()

    for field_name in field_names:
        # For "export all", always use global data (no filter)
        is_single = len(field_names) == 1
        ff  = export_filter_field  if is_single else ''
        fv  = export_filter_value  if is_single else ''
        ff2 = export_filter_field2 if is_single else ''
        fv2 = export_filter_value2 if is_single else ''
        ff3 = export_filter_field3 if is_single else ''
        fv3 = export_filter_value3 if is_single else ''

        rows = list(
            _snapshot_qs_for_request(request, field_name, ff, fv, ff2, fv2, ff3, fv3)
            .order_by('date')
            .values('date', 'counts')
        )
        if not rows:
            continue

        date_set     = sorted({r['date'] for r in rows}, reverse=True)
        value_totals: dict = {}
        data_matrix:  dict = {}

        for r in rows:
            for v, count in r['counts'].items():
                if v not in data_matrix:
                    data_matrix[v]  = {}
                    value_totals[v] = 0
                data_matrix[v][r['date']] = count
                value_totals[v] += count

        value_set = sorted(value_totals)  # alphabetical for export columns

        sheet_title = label_map.get(field_name, field_name)[:31]
        ws = wb.create_sheet(title=sheet_title)

        # ── Summary row ──────────────────────────────────────────────────────────
        if request.GET.get('date_from') or request.GET.get('date_to'):
            date_from_str = request.GET.get('date_from', '…')
            date_to_str   = request.GET.get('date_to', '…')
            period_str = f"From {date_from_str} to {date_to_str}"
        else:
            try:
                n_days = int(request.GET.get('days', 90))
            except ValueError:
                n_days = 90
            period_str = f"Last {n_days} days" if n_days > 0 else "All data"

        filter_parts = []
        if ff  and fv:  filter_parts.append(f"{ff} = {fv}")
        if ff2 and fv2: filter_parts.append(f"{ff2} = {fv2}")
        if ff3 and fv3: filter_parts.append(f"{ff3} = {fv3}")
        filter_str = " + ".join(filter_parts) if filter_parts else "No filter"

        summary = (
            f"{label_map.get(field_name, field_name)}"
            f"  |  {period_str}"
            f"  |  Filter: {filter_str}"
            f"  |  Exported {datetime.date.today():%d.%m.%Y}"
        )

        headers = ['Date', 'Total'] + [v or '(empty)' for v in value_set]
        summary_cell = ws.cell(row=1, column=1, value=summary)
        summary_cell.font      = Font(italic=True, color='5A6272')
        summary_cell.fill      = PatternFill('solid', fgColor='F0F2F7')
        summary_cell.alignment = Alignment(horizontal='left', vertical='center')
        if len(headers) > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.row_dimensions[1].height = 18

        # ── Header row ───────────────────────────────────────────────────────────
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align

        # ── Data rows ────────────────────────────────────────────────────────────
        bold_font = Font(bold=True)
        for row_idx, date in enumerate(date_set, 3):
            ws.cell(row=row_idx, column=1, value=date.strftime('%d.%m.%Y')).alignment = center_align
            row_total = sum(data_matrix[v].get(date, 0) for v in value_set)
            total_cell = ws.cell(row=row_idx, column=2, value=row_total)
            total_cell.alignment = center_align
            total_cell.font      = bold_font
            total_cell.border    = thin_border
            for col_idx, v in enumerate(value_set, 3):
                cell = ws.cell(row=row_idx, column=col_idx, value=data_matrix[v].get(date, 0))
                cell.alignment = center_align
                cell.border    = thin_border

        # Auto column widths
        for col_idx, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, len(str(h)) + 2)

        ws.freeze_panes = 'B3'

    if not wb.sheetnames:
        return HttpResponse('No data found', status=404)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── History Query (ServerHistory SCD2) ───────────────────────────────────────

@login_required
def history_query(request):
    """Render the ad-hoc history query builder page."""
    from inventory.history_query_engine import get_meta, get_all_filter_values

    meta = get_meta()
    field_values = get_all_filter_values()

    last_row  = ServerHistory.objects.order_by('-valid_from').values('valid_from').first()
    last_date = last_row['valid_from'] if last_row else None

    presets = [
        {'id': 'os-trends',   'label': 'OS trends'},
        {'id': 'cpu-region',  'label': 'CPU by Region'},
        {'id': 'srv-env',     'label': 'Servers by Env'},
        {'id': 'os-region',   'label': 'OS × Region'},
        {'id': 'ram-dc',      'label': 'RAM by Datacenter'},
        {'id': 'srv-type',    'label': 'Servers by Type'},
        {'id': 'srv-maker',   'label': 'Servers by Manufacturer'},
        {'id': 'apps-region', 'label': 'Apps by Region'},
    ]

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    saved_qs = list(
        SavedHistoryQuery.objects.filter(user_profile=profile)
        .values('id', 'name', 'spec')
    )

    return render(request, 'inventory/history_query.html', {
        'meta_json':          json.dumps(meta),
        'field_values_json':  json.dumps(field_values),
        'last_date':          last_date,
        'presets':            presets,
        'today':              datetime.date.today().isoformat(),
        'saved_queries_json': json.dumps(saved_qs),
    })


@login_required
def history_data(request):
    """Run a history query and return JSON for Chart.js."""
    from datetime import date as _date
    from inventory.history_query_engine import get_meta, run_query

    meta = get_meta()
    fields_meta = meta['fields']

    measure_field = request.GET.get('measure_field', 'SERVER_ID')
    if measure_field not in fields_meta:
        return JsonResponse({'error': 'Invalid measure field'}, status=400)

    measure_agg = request.GET.get('measure_agg', 'count_distinct')
    allowed_aggs = fields_meta[measure_field].get('aggregations', [])
    if measure_agg not in allowed_aggs:
        return JsonResponse({'error': 'Invalid aggregation for this field'}, status=400)

    group_by = [
        g for g in request.GET.getlist('group_by')
        if g in fields_meta and fields_meta[g].get('groupable')
    ][:2]

    filter_fields = request.GET.getlist('filter_field')
    filter_ops    = request.GET.getlist('filter_op')
    filter_values = request.GET.getlist('filter_value')
    filters = [
        {'field': ff, 'op': fo, 'value': fv}
        for ff, fo, fv in zip(filter_fields, filter_ops, filter_values)
        if ff in fields_meta and fields_meta[ff].get('filterable')
    ]

    try:
        start = _date.fromisoformat(request.GET.get('start', '2026-01-01'))
        end   = _date.fromisoformat(request.GET.get('end',   str(_date.today())))
    except ValueError:
        return JsonResponse({'error': 'Invalid date'}, status=400)

    step = request.GET.get('step', 'month')
    if step not in ('day', 'week', 'month'):
        step = 'month'
    chart_limit = min(int(request.GET.get('top_n', '10')), 30)

    data = run_query(
        {'measure_field': measure_field, 'measure_agg': measure_agg,
         'group_by': group_by, 'filters': filters,
         'start': start, 'end': end, 'step': step},
        chart_limit=chart_limit,
    )
    return JsonResponse(data)


@login_required
def history_export(request):
    """Export the current history query result as an Excel file (all series)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as _date
    from inventory.history_query_engine import get_meta, run_query

    meta = get_meta()
    fields_meta = meta['fields']

    measure_field = request.GET.get('measure_field', 'SERVER_ID')
    measure_agg   = request.GET.get('measure_agg', 'count_distinct')
    group_by = [g for g in request.GET.getlist('group_by') if g in fields_meta][:2]
    filter_fields = request.GET.getlist('filter_field')
    filter_ops    = request.GET.getlist('filter_op')
    filter_values = request.GET.getlist('filter_value')
    filters = [
        {'field': ff, 'op': fo, 'value': fv}
        for ff, fo, fv in zip(filter_fields, filter_ops, filter_values)
        if ff in fields_meta and fields_meta[ff].get('filterable')
    ]

    try:
        start = _date.fromisoformat(request.GET.get('start', '2026-01-01'))
        end   = _date.fromisoformat(request.GET.get('end',   str(_date.today())))
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    step = request.GET.get('step', 'month')
    # Export all series — chart_limit ignored in Excel
    data = run_query(
        {'measure_field': measure_field, 'measure_agg': measure_agg,
         'group_by': group_by, 'filters': filters,
         'start': start, 'end': end, 'step': step},
        chart_limit=9999,
    )

    measure_label = fields_meta.get(measure_field, {}).get('label', measure_field)
    agg_label     = meta.get('agg_labels', {}).get(measure_agg, measure_agg)
    gb_labels     = ' + '.join(fields_meta.get(g, {}).get('label', g) for g in group_by) or 'None'
    filter_desc   = ', '.join(f"{f['field']}={f['value']}" for f in filters) or 'None'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'History Query'

    summary = (
        f"Measure: {measure_label} ({agg_label}) | Group by: {gb_labels} | "
        f"Filters: {filter_desc} | Period: {start} → {end} ({step}) | "
        f"Exported {_date.today().strftime('%d.%m.%Y')}"
    )
    ws.append([summary])
    ws.cell(1, 1).font = Font(italic=True, size=10)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(data['datasets']) + 1))

    true_totals = data.get('true_totals')
    headers = ['Date', 'Total'] + [ds['label'] for ds in data['datasets']]
    ws.append(headers)
    hdr_fill = PatternFill('solid', fgColor='A8D5BA')
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(2, col_idx)
        cell.font = Font(bold=True, size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, len(str(h)) + 2)

    ws.freeze_panes = 'C3'

    for i, label in enumerate(data['labels']):
        total_val = true_totals[i] if true_totals else sum(ds['data'][i] or 0 for ds in data['datasets'])
        row = [label, total_val] + [ds['data'][i] for ds in data['datasets']]
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"server_history_{measure_field}_{_date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def history_drilldown(request):
    """Server list for a specific date+group combination (right panel drill-down)."""
    from datetime import date as _date
    from inventory.history_query_engine import get_meta, drilldown_servers

    meta = get_meta()
    fields_meta = meta['fields']

    try:
        point_date = _date.fromisoformat(request.GET.get('date', ''))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date'}, status=400)

    group_by     = [g for g in request.GET.getlist('group_by') if g in fields_meta][:2]
    group_values = request.GET.getlist('group_value')[:2]

    measure_field = request.GET.get('measure_field', 'SERVER_ID')
    if measure_field not in fields_meta:
        measure_field = 'SERVER_ID'
    measure_agg = request.GET.get('measure_agg', 'count_distinct')
    if measure_agg not in ('count_distinct', 'sum'):
        measure_agg = 'count_distinct'

    filter_fields = request.GET.getlist('filter_field')
    filter_ops    = request.GET.getlist('filter_op')
    filter_vals   = request.GET.getlist('filter_value')
    filters = [
        {'field': f, 'op': o, 'value': v}
        for f, o, v in zip(filter_fields, filter_ops, filter_vals)
        if f in fields_meta and fields_meta[f].get('filterable')
    ]

    result = drilldown_servers(point_date, group_by, group_values, filters,
                               measure_field=measure_field, measure_agg=measure_agg)
    return JsonResponse(result)


@login_required
def save_history_query(request):
    """Save or overwrite a named history query for the current user."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    import json as _json
    try:
        body = _json.loads(request.body)
    except ValueError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    name = (body.get('name') or '').strip()
    spec = body.get('spec')
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    if len(name) > 100:
        return JsonResponse({'error': 'Name must be 100 characters or fewer'}, status=400)
    if not spec or not isinstance(spec, dict):
        return JsonResponse({'error': 'Spec is required'}, status=400)

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    obj, created = SavedHistoryQuery.objects.update_or_create(
        user_profile=profile,
        name=name,
        defaults={'spec': spec},
    )
    return JsonResponse({'id': obj.id, 'name': obj.name, 'created': created})


@login_required
def delete_history_query(request, query_id):
    """Delete a saved history query owned by the current user."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    deleted, _ = SavedHistoryQuery.objects.filter(id=query_id, user_profile=profile).delete()
    if not deleted:
        return JsonResponse({'error': 'Not found'}, status=404)
    return JsonResponse({'deleted': query_id})

