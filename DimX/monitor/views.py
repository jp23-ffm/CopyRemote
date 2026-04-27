import io
import socket
from datetime import date, timedelta

from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views import View

from monitor.models import (
    AuditConnection,
    GlobalHealthStatus,
    HostHealthStatus,
    StatsConcurrentUsers,
    StatsRequest,
)

User = get_user_model()


"""class ClusterStatusView(View):
    # View to read the statuts of all hosts + the global checks from the DB
    
    STALE_THRESHOLD_SECONDS = 600  # 10 minutes for the hosts
    GLOBAL_STALE_THRESHOLD_SECONDS = 18000  # 5 hours for the global checks
    
    def get(self, request, *args, **kwargs):
        hosts = HostHealthStatus.objects.all().order_by('host_name')
        
        # Get the global checks
        try:
            global_health = GlobalHealthStatus.get_or_create_singleton()
        except:
            global_health = None
        
        if not hosts.exists() and not global_health:
            return JsonResponse({
                "cluster_status": "Unknown",
                "message": "No reported status yet",
                "timestamp": timezone.now().isoformat(),
                "hosts": [],
                "global_checks": None
            }, status=503)
        
        # ===== hosts =====
        hosts_data = []
        cluster_status = "OK"
        stale_hosts = []
        error_hosts = []
        warning_hosts = []
        
        for host in hosts:
            staleness = host.get_staleness_seconds()
            is_stale = host.is_stale(self.STALE_THRESHOLD_SECONDS)
            
            if is_stale:
                effective_status = "Stale"
                stale_hosts.append(host.host_name)
            else:
                effective_status = host.status
                if host.status == "Error":
                    error_hosts.append(host.host_name)
                elif host.status == "Warning":
                    warning_hosts.append(host.host_name)
            
            host_info = {
                "host_name": host.host_name,
                "status": effective_status,
                "reported_status": host.status,
                "last_updated": host.last_updated.isoformat() if host.last_updated else None,
                "staleness_seconds": round(staleness, 2) if staleness else None,
                "is_stale": is_stale,
                "hostname": host.hostname,
                "ip_address": host.ip_address,
                "version": host.version,
                "uptime": host.uptime,
                "checks_summary": {
                    "total": host.checks_data.get('total_checks', 0),
                    "ok": host.checks_data.get('ok_count', 0),
                    "warning": host.checks_data.get('warning_count', 0),
                    "error": host.checks_data.get('error_count', 0),
                },
                "checks": host.checks_data.get('checks', [])
            }
            
            hosts_data.append(host_info)
        
        # Statut des hosts
        if stale_hosts:
            cluster_status = "Degraded" if len(stale_hosts) < hosts.count() else "Critical"
        elif error_hosts:
            cluster_status = "Degraded" if len(error_hosts) < hosts.count() else "Critical"
        elif warning_hosts:
            cluster_status = "Warning"
        else:
            cluster_status = "OK"
        
        # ===== GLOBAL CHECKS =====
        global_checks_data = None
        if global_health:
            global_staleness = global_health.get_staleness_seconds()
            global_is_stale = global_health.is_stale(self.GLOBAL_STALE_THRESHOLD_SECONDS)
            
            global_checks_data = {
                "status": "Stale" if global_is_stale else global_health.status,
                "reported_status": global_health.status,
                "last_updated": global_health.last_updated.isoformat() if global_health.last_updated else None,
                "staleness_seconds": round(global_staleness, 2) if global_staleness else None,
                "is_stale": global_is_stale,
                "checks_summary": {
                    "total": global_health.checks_data.get('total_checks', 0),
                    "ok": global_health.checks_data.get('ok_count', 0),
                    "warning": global_health.checks_data.get('warning_count', 0),
                    "error": global_health.checks_data.get('error_count', 0),
                },
                "checks": global_health.checks_data.get('checks', [])
            }
            
            # Take account of the global check for the whole status
            if global_is_stale:
                if cluster_status == "OK":
                    cluster_status = "Warning"
            elif global_health.status == "Error":
                if cluster_status in ["OK", "Warning"]:
                    cluster_status = "Degraded"
            elif global_health.status == "Warning":
                if cluster_status == "OK":
                    cluster_status = "Warning"
        
        # ===== RESPONSE =====
        response = {
            "cluster_status": cluster_status,
            "timestamp": timezone.now().isoformat(),
            "summary": {
                "total_hosts": hosts.count(),
                "healthy_hosts": hosts.count() - len(stale_hosts) - len(error_hosts),
                "stale_hosts": len(stale_hosts),
                "error_hosts": len(error_hosts),
                "warning_hosts": len(warning_hosts),
            },
            "hosts": hosts_data,
            "global_checks": global_checks_data
        }
        
        # Issues
        issues = []
        if stale_hosts:
            issues.append({
                "type": "stale",
                "message": f"host(s) haven't reported since {self.STALE_THRESHOLD_SECONDS}s",
                "affected_hosts": stale_hosts
            })
        if error_hosts:
            issues.append({
                "type": "error",
                "message": "host(s) on error",
                "affected_hosts": error_hosts
            })
        if global_checks_data and global_checks_data["is_stale"]:
            issues.append({
                "type": "global_stale",
                "message": f"Global checks not updated since {self.GLOBAL_STALE_THRESHOLD_SECONDS}s",
                "affected": "global_checks"
            })
        if global_checks_data and global_checks_data["status"] == "Error":
            issues.append({
                "type": "global_error",
                "message": "Global checks on error",
                "affected": "global_checks"
            })
        
        if issues:
            response["issues"] = issues
        
        http_status = 503 if cluster_status in ["Critical", "Degraded"] else 200

        return JsonResponse(response, status=http_status, json_dumps_params={'indent': 2})
"""

def get_cluster_status(simple=False):
    hosts = HostHealthStatus.objects.all().order_by('host_name')
    STALE_THRESHOLD_SECONDS = 600  # 10 minutes for the hosts
    GLOBAL_STALE_THRESHOLD_SECONDS = 7200  # 2 hours for the global checks

    # Get the global checks
    try:
        global_health = GlobalHealthStatus.get_or_create_singleton()
    except:
        global_health = None

    if not hosts.exists() and not global_health:
        return {
            "cluster_status": "Unknown",
            "message": "No reported status yet",
            "timestamp": timezone.now().isoformat()
        }, 503

    hostname = socket.gethostname()

    # ===== hosts =====
    cluster_status = "OK"
    stale_hosts = []
    error_hosts = []
    warning_hosts = []

    for host in hosts:
        staleness = host.get_staleness_seconds()
        is_stale = host.is_stale(STALE_THRESHOLD_SECONDS)

        if is_stale:
            stale_hosts.append(host.host_name)
        else:
            if host.status == "Error":
                error_hosts.append(host.host_name)
            elif host.status == "Warning":
                warning_hosts.append(host.host_name)

    # Statut des hosts
    if stale_hosts:
        cluster_status = "Degraded" if len(stale_hosts) < hosts.count() else "Critical"
    elif error_hosts:
        cluster_status = "Degraded" if len(error_hosts) < hosts.count() else "Critical"
    elif warning_hosts:
        cluster_status = "Warning"
    else:
        cluster_status = "OK"

    # ===== GLOBAL CHECKS =====
    if global_health:
        global_staleness = global_health.get_staleness_seconds()
        global_is_stale = global_health.is_stale(GLOBAL_STALE_THRESHOLD_SECONDS)

        # Take account of the global check for the whole status
        if global_is_stale:
            if cluster_status == "OK":
                cluster_status = "Warning"
        elif global_health.status == "Error":
            if cluster_status in ["OK", "Warning"]:
                cluster_status = "Degraded"
        elif global_health.status == "Warning":
            if cluster_status == "OK":
                cluster_status = "Warning"

    if simple:
        return {
            "cluster_status": cluster_status,
            "timestamp": timezone.now().isoformat()
        }, 200

    # ===== RESPONSE =====
    response = {
        "cluster_status": cluster_status,
        "hostname": hostname,
        "timestamp": timezone.now().isoformat(),
        "summary": {
            "total_hosts": hosts.count(),
            "healthy_hosts": hosts.count() - len(stale_hosts) - len(error_hosts),
            "stale_hosts": len(stale_hosts),
            "error_hosts": len(error_hosts),
            "warning_hosts": len(warning_hosts),
        },
        "hosts": [
            {
                "host_name": host.host_name,
                "status": "Stale" if host.is_stale(600) else host.status,
                "reported_status": host.status,
                "last_updated": host.last_updated.isoformat() if host.last_updated else None,
                "staleness_seconds": round(host.get_staleness_seconds(), 2) if host.get_staleness_seconds() else None,
                "is_stale": host.is_stale(600),
                "hostname": host.hostname,
                "ip_address": host.ip_address,
                "version": host.version,
                "uptime": host.uptime,
                "checks_summary": {
                    "total": host.checks_data.get('total_checks', 0),
                    "ok": host.checks_data.get('ok_count', 0),
                    "warning": host.checks_data.get('warning_count', 0),
                    "error": host.checks_data.get('error_count', 0),
                },
                "checks": host.checks_data.get('checks', [])
            } for host in hosts
        ],
        "global_checks": {
            "status": "Stale" if global_is_stale else global_health.status,
            "reported_status": global_health.status,
            "last_updated": global_health.last_updated.isoformat() if global_health.last_updated else None,
            "staleness_seconds": round(global_staleness, 2) if global_staleness else None,
            "is_stale": global_is_stale,
            "checks_summary": {
                "total": global_health.checks_data.get('total_checks', 0),
                "ok": global_health.checks_data.get('ok_count', 0),
                "warning": global_health.checks_data.get('warning_count', 0),
                "error": global_health.checks_data.get('error_count', 0),
            },
            "checks": global_health.checks_data.get('checks', [])
        } if global_health else None
    }

    # Issues
    issues = []
    if stale_hosts:
        issues.append({
            "type": "stale",
            "message": f"host(s) n'ont pas reporté depuis 600s",
            "affected_hosts": stale_hosts
        })
    if error_hosts:
        issues.append({
            "type": "error",
            "message": "host(s) en erreur",
            "affected_hosts": error_hosts
        })
    if global_health and global_is_stale:
        issues.append({
            "type": "global_stale",
            "message": f"Checks globaux non mis à jour depuis 7200s",
            "affected": "global_checks"
        })
    if global_health and global_health.status == "Error":
        issues.append({
            "type": "global_error",
            "message": "Checks globaux en erreur",
            "affected": "global_checks"
        })

    if issues:
        response["issues"] = issues

    http_status = 503 if cluster_status in ["Critical", "Degraded"] else 200

    return response, http_status


# ─── Stats dashboard ──────────────────────────────────────────────────────────

def stats_dashboard(request):
    apps = list(
        StatsRequest.objects.values_list('app_name', flat=True)
        .distinct()
        .order_by('app_name')
    )
    users = list(
        User.objects.filter(stats_requests__isnull=False)
        .distinct()
        .values('id', 'username')
        .order_by('username')
    )
    return render(request, 'monitor/stats_dashboard.html', {
        'apps': apps,
        'users': users,
    })


def _parse_days(request, default=7, max_days=365):
    try:
        days = int(request.GET.get('days', default))
        return min(max(1, days), max_days)
    except (ValueError, TypeError):
        return default


def stats_top_views(request):
    """Bar chart: top N views by total hits over the last N days."""
    days = _parse_days(request)
    app = request.GET.get('app', '')
    user_id = request.GET.get('user_id', '')
    limit = min(int(request.GET.get('limit', 15)), 50)

    since = date.today() - timedelta(days=days)
    qs = StatsRequest.objects.filter(date__gte=since)
    if app:
        qs = qs.filter(app_name=app)
    if user_id:
        qs = qs.filter(user_id=user_id)

    from django.db.models import Sum
    rows = (
        qs.values('view_name', 'app_name')
        .annotate(total=Sum('hit_count'))
        .order_by('-total')[:limit]
    )
    return JsonResponse({
        'labels': [r['view_name'] for r in rows],
        'values': [r['total'] for r in rows],
        'apps': [r['app_name'] for r in rows],
    })


def stats_hits_by_day(request):
    """Line chart: total hits per day, optionally filtered by app/user."""
    days = _parse_days(request, default=30)
    app = request.GET.get('app', '')
    user_id = request.GET.get('user_id', '')

    since = date.today() - timedelta(days=days)
    qs = StatsRequest.objects.filter(date__gte=since)
    if app:
        qs = qs.filter(app_name=app)
    if user_id:
        qs = qs.filter(user_id=user_id)

    rows = (
        qs.values('date')
        .annotate(total=Sum('hit_count'))
        .order_by('date')
    )
    return JsonResponse({
        'labels': [str(r['date']) for r in rows],
        'values': [r['total'] for r in rows],
    })


def stats_hits_by_app(request):
    """Doughnut chart: total hits per app over the last N days."""
    days = _parse_days(request)
    user_id = request.GET.get('user_id', '')

    since = date.today() - timedelta(days=days)
    qs = StatsRequest.objects.filter(date__gte=since)
    if user_id:
        qs = qs.filter(user_id=user_id)

    rows = (
        qs.values('app_name')
        .annotate(total=Sum('hit_count'))
        .order_by('-total')
    )
    return JsonResponse({
        'labels': [r['app_name'] for r in rows],
        'values': [r['total'] for r in rows],
    })


def stats_concurrent(request):
    """Line chart: active session snapshots over the last N days."""
    days = _parse_days(request, default=3)
    since = timezone.now() - timedelta(days=days)
    rows = (
        StatsConcurrentUsers.objects
        .filter(snapshot_at__gte=since, app_name__isnull=True)
        .order_by('snapshot_at')
        .values('snapshot_at', 'active_users')
    )
    return JsonResponse({
        'labels': [r['snapshot_at'].isoformat() for r in rows],
        'values': [r['active_users'] for r in rows],
    })


def stats_connections(request):
    """Table data: recent auth events (login / logout / login_failed)."""
    days = _parse_days(request, default=1)
    action = request.GET.get('action', '')
    user_id = request.GET.get('user_id', '')
    limit = min(int(request.GET.get('limit', 100)), 500)

    since = timezone.now() - timedelta(days=days)
    qs = AuditConnection.objects.filter(timestamp__gte=since)
    if action:
        qs = qs.filter(action=action)
    if user_id:
        qs = qs.filter(user_id=user_id)

    rows = qs.order_by('-timestamp').values(
        'username', 'action', 'ip_address', 'timestamp'
    )[:limit]

    return JsonResponse({
        'rows': [
            {
                'username': r['username'],
                'action': r['action'],
                'ip': r['ip_address'] or '',
                'timestamp': r['timestamp'].isoformat(),
            }
            for r in rows
        ]
    })


def stats_export_xlsx(request):
    """Export stats tables as a multi-sheet XLSX file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    days = _parse_days(request, default=30, max_days=365)
    app = request.GET.get('app', '')
    user_id = request.GET.get('user_id', '')
    since_date = date.today() - timedelta(days=days)
    since_dt = timezone.now() - timedelta(days=days)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_align = Alignment(horizontal='center')

    def _style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = 'A2'

    def _autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    # ── Sheet 1 : stats_requests ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'View hits'
    qs1 = StatsRequest.objects.filter(date__gte=since_date)
    if app:
        qs1 = qs1.filter(app_name=app)
    if user_id:
        qs1 = qs1.filter(user_id=user_id)

    headers1 = ['Date', 'App', 'View', 'Method', 'User', 'Hits']
    _style_header(ws1, headers1)
    for r in qs1.select_related('user').order_by('-date', '-hit_count'):
        ws1.append([
            r.date.isoformat(),
            r.app_name,
            r.view_name,
            r.method,
            r.user.username if r.user else '',
            r.hit_count,
        ])
    _autowidth(ws1)

    # ── Sheet 2 : audit_connections ──────────────────────────────────────────
    ws2 = wb.create_sheet('Auth events')
    qs2 = AuditConnection.objects.filter(timestamp__gte=since_dt)
    if user_id:
        qs2 = qs2.filter(user_id=user_id)

    headers2 = ['Timestamp', 'Username', 'Action', 'IP address']
    _style_header(ws2, headers2)
    for r in qs2.order_by('-timestamp'):
        ws2.append([
            r.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            r.username,
            r.action,
            r.ip_address or '',
        ])
    _autowidth(ws2)

    # ── Sheet 3 : stats_concurrent_users ─────────────────────────────────────
    ws3 = wb.create_sheet('Concurrent users')
    qs3 = StatsConcurrentUsers.objects.filter(snapshot_at__gte=since_dt, app_name__isnull=True)
    headers3 = ['Snapshot at', 'Active sessions']
    _style_header(ws3, headers3)
    for r in qs3.order_by('snapshot_at'):
        ws3.append([
            r.snapshot_at.strftime('%Y-%m-%d %H:%M:%S'),
            r.active_users,
        ])
    _autowidth(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"chimera_stats_{date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
